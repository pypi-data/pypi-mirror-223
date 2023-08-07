from sqlalchemy import create_engine, MetaData
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
from celery import Celery

from xj_invoice.services.invoice_service import InvoiceService


class ExeclImportService:
    @staticmethod
    def execl_import(execl_file, module):
        old_keys = []
        new_key = ''
        if module == "invoice":
            # 预先定义的数据映射
            data_mappings = {
                "开票时间": "invoice_time",
                "发票号码": "invoice_number",
                "开票金额": "invoice_price",
                "税率": "tax_rate",
                "检索发票税额": "invoice_tax",
                "专/普": "invoice_type",
                "是否作废": "invoice_status",
                "检索发票金额（不含税）": "amount_excluding_tax"
            }
        elif module == "thread":
            return "合同模板", None
        elif module == "finance":
            return "财务模板", None
        else:
            return None, "未定义导入模板"
        # 读取 Excel 文件
        wb = openpyxl.load_workbook(execl_file)
        sheet = wb.active

        # 拆分合并单元格并提取数据
        merged_cells_ranges = list(sheet.merged_cells.ranges)
        for merged_cell in merged_cells_ranges:
            min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
            value = sheet.cell(row=min_row, column=min_col).value
            range_string = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            sheet.unmerge_cells(range_string)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = value

        # 从 Excel 中获取数据并创建 DataFrame
        data = list(sheet.values)
        df = pd.DataFrame(data[1:], columns=data[0])

        # 数据预处理
        df = df.astype(str)
        df.replace('nan', '', inplace=True)

        # 对 DataFrame 的每一行应用数据映射，生成对应的数据库插入数据
        db_data = []
        # for _, row in df.iterrows():
        #     row_data_all = {data_mappings[k]: v for k, v in row.to_dict().items() if k in data_mappings}
        #     # 首先添加包含所有字段的字典
        #     db_data.append({k: row_data_all[k] for k in ["invoice_time", "invoice_number", "invoice_price"]})
        #     # 对于需要单独成字典的字段，分别创建新的字典并添加到列表中
        #     db_data.append({"tax_rate": row_data_all["tax_rate"]})
        #     db_data.append({"invoice_tax": row_data_all["invoice_tax"]})
        for _, row in df.iterrows():
            row_data_db = {data_mappings[k]: v for k, v in row.to_dict().items() if k in data_mappings}
            db_data.append(row_data_db)

        # db_data 就是你要的列表，你可以用它来插入数据到数据库
        for item in db_data:
            if module == "invoice":
                if item['invoice_type'] == "普":
                    item['invoice_type_code'] = "GENERAL_VAT_INVOICE"
                elif item['invoice_type'] == "专":
                    item['invoice_type_code'] = "SPECIAL_VAT_INVOICE"

                if item['invoice_status'] == "是":
                    item['invoice_status_code'] = "CANCEL"
                else:
                    item['invoice_status_code'] = "INVOIC"

                item['thread_id'] = 2479
                item['user_id'] = 686
                invoice, invoice_err = InvoiceService.add(item)
            if invoice_err:
                print(">>>>>", invoice_err)
            print(invoice.get("id", 0))

        return None, None
