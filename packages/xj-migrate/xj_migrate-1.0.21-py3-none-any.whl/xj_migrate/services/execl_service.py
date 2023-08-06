import random
import os
import re
import traceback
import logging
import uuid
from uuid import uuid4
import numpy as np
import redis
from django.core.cache import cache
from openpyxl.reader.excel import load_workbook

from main.settings import BASE_DIR
from pathlib import Path
from django.forms.models import model_to_dict
from sqlalchemy import create_engine, Table, MetaData, exc, insert, text, update
import pandas as pd
import openpyxl
import datetime
from openpyxl.utils import get_column_letter

from xj_finance.services.finance_transact_service import FinanceTransactService
from xj_role.services.user_group_service import UserGroupService
from xj_thread.services.thread_category_service import ThreadCategoryService
from xj_thread.services.thread_item_service import ThreadItemService
from xj_thread.services.thread_list_service import ThreadListService
from xj_user.services.user_detail_info_service import DetailInfoService
from xj_user.services.user_service import UserService
from .execl_import import ExeclImportService
from ..models import MigrateExeclConfig, MigrateImportProgress
from ..utils.execl_import import execl_import
# from ..utils.execl_import import execl_import
from ..utils.j_config import JConfig
from ..utils.j_dict import JDict
from celery import shared_task, current_task
import redis

from xj_invoice.services.invoice_service import InvoiceService
from ..utils.utility_method import replace_key_in_list_dicts, replace_key_in_list_replacement_dicts, \
    replace_key_in_dict_replacement_dicts

logger = logging.getLogger(__name__)

module_root = str(Path(__file__).resolve().parent)
# 配置之对象
main_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="main"))
module_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="main"))

import_main_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="xj_migrate"))
import_module_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="xj_migrate"))

# 获取数据库信息
user = main_config_dict.mysql_user or module_config_dict.mysql_user or ""
password = main_config_dict.mysql_password or module_config_dict.mysql_password or ""
host = main_config_dict.mysql_host or module_config_dict.mysql_host or ""
port = main_config_dict.mysql_port or module_config_dict.mysql_port or ""
database = main_config_dict.mysql_database or module_config_dict.mysql_database or ""

import_logs = import_main_config_dict.import_logs or import_module_config_dict.import_logs or ""


def unique_identification():
    return "HL-" + "ZM" + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + "-" + str(random.randint(10000, 99999))


CELERY_TRACK_STARTED = True

# 沙盒
INVOICE_RECEIVABLE = "INVOICE_RECEIVABLE"  # 发票-应收
MANAGEMENT_FEE_RECEIVABLE = "MANAGEMENT_FEE_RECEIVABLE"  # 管理费-应收
TAX_RECEIVABLES = "TAX_RECEIVABLES"  # 税金-应收
COMMISSION_RECEIVABLE = "COMMISSION_RECEIVABLE"  # 佣金-应收
DEDUCTION_BASIS = "DEDUCTION_BASIS"  # 扣款依据


class ExeclService:
    @shared_task(bind=True)
    def bind_data_processing(self, execl_file, module, group_id):
        task_id = str(uuid.uuid4())
        if not group_id:
            return None, "所属企业不能为空"
        group, group_err = UserGroupService.details_group({"id": group_id})
        if group_err:
            return None, group_err
        if group:
            their_user_name, err = DetailInfoService.get_detail(
                search_params={"full_name": group.get("description", "")})
            if their_user_name:
                their_user_id = their_user_name['user_id']
            else:
                # 如果不存在该客户创建
                user, user_err = UserService.user_add(
                    {"full_name": group.get("description", ""), "user_type": "UNKNOWN"})
                if user_err:
                    return None, user_err
                their_user_id = user['user_id']
                # 创建成功绑定用户企业关系
                user_group, user_group_err = UserGroupService.user_bind_groups(their_user_id, group_id)
                if user_group_err:
                    return None, user_group_err

        category, category_err = ThreadCategoryService.list(params={"value": "LABOR_CONTRACT"})
        if not category['list'] and category_err:
            return None, category_err
        category_id = category['list'][0]['id']
        try:
            filepath = ""
            relate_uuid = ""
            success_count = 0
            failure_count = 0
            failures = []
            replacement_dict = {}
            unique_id_row_indices = {}
            remain_fields_values = {}
            if module == "invoice":
                # 预先定义的数据映射
                data_mappings = {
                    "客户新编码": {"db_field": "order_no", "type": "remain", "default": ""},
                    "开票时间": {"db_field": "invoice_time", "type": "remain", "default": ""},
                    "发票号码": {"db_field": "invoice_number", "type": "remain", "default": ""},
                    "开票金额": {"db_field": "invoice_price", "type": "remain", "default": 0},
                    "税率": {"db_field": "tax_rate", "type": "remain", "default": ""},
                    "检索发票税额": {"db_field": "invoice_tax", "type": "remain", "default": ""},
                    "专/普": {"db_field": "invoice_type", "type": "remain", "default": ""},
                    "是否作废": {"db_field": "invoice_status", "type": "remain", "default": ""}
                }
            elif module == "user":
                data_mappings = {
                    "客户名称": {"db_field": "full_name", "type": "remain"},
                    "客户类型": {"db_field": "user_type", "type": "remain"},
                }
            elif module == "thread":
                data_mappings = {
                    "归属地区": {"db_field": "belonging_region", "type": "remain", "default": ""},
                    "业务员": {"db_field": "salesman", "type": "remain", "default": ""},
                    "收回时间": {"db_field": "withdrawal_date", "type": "remain", "default": ""},
                    "客户新编码": {"db_field": "customer_code", "type": "remain", "default": ""},
                    "客户名称": {"db_field": "full_name", "type": "remain", "default": ""},
                    "工程项目": {"db_field": "title", "type": "remain", "default": ""},
                    "工程地址": {"db_field": "project_address", "type": "remain", "default": ""},
                    # "开工日期": {"db_field": "commencement_date", "type": "remain", "default": ""},
                    # "完工日期": {"db_field": "completion_date", "type": "remain", "default": ""},
                    "合同编码": {"db_field": "contract_code", "type": "remain", "default": ""},
                    "合同金额": {"db_field": "contract_amount", "type": "remain", "default": 0},
                    "分包合同": {"db_field": "subcontracts", "type": "remain", "default": ""},
                }
            elif module == "finance":

                replacement_dict = {
                    'remittance_amount': "amount",
                    'management_fees': "amount",
                    'taxes': "amount",
                    'commission': "amount",
                    'amount_remitted': "amount",
                    'deduction_basis': "amount",
                    "remit_time": "transact_time"
                }

                data_mappings = {
                    "客户新编码": {"db_field": "order_no", "type": "remain", "default": ""},
                    "摘要": {"db_field": "summary", "type": "remain", "default": ""},
                    "备注1": {"db_field": "remark", "type": "remain", "default": ""},
                    "汇入时间": {"db_field": "transact_time", "type": "remain", "default": ""},
                    "汇入金额": {"db_field": "remittance_amount", "type": "remain", "default": 0,
                                 "additional_value": {"master_data": "YES"}},
                    "扣费依据": {"db_field": "deduction_basis", "type": "extract", "default": 0,
                                 "additional_value": {"sand_box": "DEDUCTION_BASIS", "order_no": "", "summary": "",
                                                      "transact_time": ""}},
                    "管理费点数": {"db_field": "manage_point", "type": "remain"},
                    "管理费": {"db_field": "management_fees", "type": "extract",
                               "additional_value": {"order_no": "", "manage_point": "", "summary": "",
                                                    "sand_box": "MANAGEMENT_FEE_RECEIVABLE"}, "negate": True},
                    "税金点数": {"db_field": "tax_point", "type": "remain"},
                    "税金": {"db_field": "taxes", "type": "extract", "default": 0,
                             "additional_value": {"sand_box": "TAX_RECEIVABLES", "order_no": "", "tax_point": "",
                                                  "summary": ""}, "negate": True},
                    "佣金点数": {"db_field": "brokerage_point", "type": "remain"},
                    "佣金": {"db_field": "commission", "type": "extract", "default": 0,
                             "additional_value": {"order_no": "", "sand_box": "COMMISSION_RECEIVABLE",
                                                  "brokerage_point": "", "summary": ""}, "negate": True},
                    "汇出时间": {"db_field": "remit_time", "type": "remain", "default": ""},
                    "汇出金额": {"db_field": "amount_remitted", "type": "extract", "default": 0,
                                 "additional_value": {"order_no": "", "remit_time": "", "summary": ""},
                                 "negate": True},
                }
            else:
                return None, "未定义导入模板"
            # 读取 Excel 文件
            wb = openpyxl.load_workbook(execl_file)
            sheet = wb.active
            # filename = os.path.basename(execl_file)
            filename = execl_file.name
            # 拆分合并单元格并提取数据
            merged_cells_ranges = list(sheet.merged_cells.ranges)
            for merged_cell in merged_cells_ranges:
                min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
                value = sheet.cell(row=min_row, column=min_col).value
                range_string = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                sheet.unmerge_cells(range_string)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        sheet.cell(row=row, column=col).value = value

            # 从 Excel 中获取数据并创建 DataFrame
            data = list(sheet.values)
            df = pd.DataFrame(data[1:], columns=data[0])

            # 数据预处理
            df = df.astype(str)
            df.replace('nan', '', inplace=True)

            # 对 DataFrame 的每一行应用数据映射，生成对应的数据库插入数据
            db_data = []
            # 设置默认值
            for column, mapping in data_mappings.items():
                if "default" in mapping:
                    df[column].replace({'': np.nan}, inplace=True)
                    df[column].fillna(mapping["default"], inplace=True)

            for _, row in df.iterrows():
                row_uuid = uuid4()  # 为每一行生成一个新的relate_uuid
                row_data_remain = {}
                extracted_data = []

                extract_index = 0  # 初始化extracted_data的索引计数器

                # 步骤1：处理所有字段
                for k, v in row.to_dict().items():
                    if k in data_mappings:
                        if k in data_mappings:
                            if v == '' or v == 'None' or v == "NaT":
                                if "default" in data_mappings[k]:
                                    v = data_mappings[k]["default"]
                            if v != 'None':
                                if "negate" in data_mappings[k] and data_mappings[k]["negate"]:
                                    try:
                                        if float(v) != 0:  # 忽略为0的值
                                            v = str(-float(v))  # 尝试转换为浮点数并取负
                                    except ValueError:
                                        pass  # 如果转换失败，保持原样
                        if data_mappings[k]["type"] == "remain":
                            row_data_remain[data_mappings[k]["db_field"]] = v
                            remain_fields_values[data_mappings[k]["db_field"]] = v

                        elif data_mappings[k]["type"] == "extract":
                            extracted_dict = {data_mappings[k]["db_field"]: v}
                            extracted_dict['unique_identification'] = row_uuid
                            extracted_data.append(extracted_dict)
                            extract_index += 1  # 增加每个提取字段的索引计数器

                row_data_remain['unique_identification'] = row_uuid

                # 步骤2：处理“保留”类型字段的附加值
                for k, mapping in data_mappings.items():
                    if mapping["type"] == "remain" and "additional_value" in mapping:
                        additional_value = mapping["additional_value"].copy()
                        for field in additional_value.keys():
                            if field in remain_fields_values:
                                additional_value[field] = remain_fields_values[field]
                        row_data_remain = {**row_data_remain,
                                           **additional_value}  # 合并row_data_remain和additional_value

                # 步骤3：处理“提取”类型字段的附加值
                extract_index = 0  # 在下一个循环之前重置索引计数器
                for k, mapping in data_mappings.items():
                    if mapping["type"] == "extract" and "additional_value" in mapping:
                        additional_value = mapping[
                            "additional_value"].copy()  # 创建additionalvalue的副本以避免修改原始字典
                        for field in additional_value.keys():
                            if field in remain_fields_values:
                                additional_value[field] = remain_fields_values[field]
                        extracted_data[extract_index] = {**extracted_data[extract_index],
                                                         **additional_value}  # 合并extracted_dict和additional_value
                        extract_index += 1  # 增加每个提取字段的索引计数器

                db_data.append(row_data_remain)
                db_data.extend(extracted_data)

            db_data = replace_key_in_list_replacement_dicts(db_data, replacement_dict)
            # db_data 就是你要的列表，你可以用它来插入数据到数据库
            counter = 0
            total = len(db_data)
            for index, row in enumerate(db_data):
                task, created = MigrateImportProgress.objects.get_or_create(task_id=task_id,
                                                                            defaults={'filename': filename,
                                                                                      'status': 'PENDING'})
                if not created:
                    task.progress = int(100 * index / total)  # 进度为已处理的行数占总行数的百分比
                    task.status = 'PROCESSING'
                    task.save()

                try:
                    counter += 1
                    if module == "invoice":
                        row['amount'] = row['invoice_price']
                        if row['invoice_type'] == "普":
                            row['invoice_type_code'] = "GENERAL_VAT_INVOICE"
                        elif row['invoice_type'] == "专":
                            row['invoice_type_code'] = "SPECIAL_VAT_INVOICE"

                        if row['invoice_status'] == "是":
                            row['invoice_status_code'] = "CANCEL"
                        else:
                            row['invoice_status_code'] = "INVOIC"
                        thread_detail, thread_err_detail = ThreadItemService.detail(
                            search_params={"customer_code": row.get("order_no", ""), "category_id": category_id},
                            sort="id")
                        if thread_detail:
                            row['thread_id'] = thread_detail['thread_id']
                            row['user_id'] = thread_detail['user_id']
                            row['their_account_id'] = thread_detail['user_id']
                        else:
                            failures.append({
                                'index': index,
                                'error': f"合同不存在 {row['customer_code']}",
                                'data': dict(row)
                            })
                            continue
                        row['account_id'] = their_user_id
                        row['their_account_id'] = thread_detail['user_id']
                        row['sand_box'] = "INVOICE_RECEIVABLE"

                        invoice_exist, invoice_exist_err = InvoiceService.detail(
                            {"invoice_number": row.get("invoice_number", "")})
                        if invoice_exist:
                            invoice, invoice_err = InvoiceService.edit(row, invoice_exist.get("invoice_id"))
                        else:
                            invoice, invoice_err = InvoiceService.add(row)
                        if invoice:
                            finance, finance_err = FinanceTransactService.add(row)
                            if finance_err:
                                failures.append({
                                    'index': index,
                                    'error': f"财务添加失败 {finance_err}",
                                    'data': dict(row)
                                })
                                continue
                        else:
                            failures.append({
                                'index': index,
                                'error': f"发票添加失败 {invoice_err}",
                                'data': dict(row)
                            })
                            continue
                        success_count += 1
                    elif module == "finance":
                        thread_detail, thread_err_detail = ThreadItemService.detail(
                            search_params={"customer_code": row.get("order_no", ""), "category_id": category_id},
                            sort="id")
                        if thread_detail:
                            row['thread_id'] = thread_detail['thread_id']
                        else:
                            failures.append({
                                'index': index,
                                'error': f"合同不存在 {row['order_no']}",
                                'data': dict(row)
                            })
                            continue
                        row['account_id'] = their_user_id
                        row['their_account_id'] = thread_detail['user_id']
                        if not row['amount']:
                            continue
                        unique_id = row.get('unique_identification')
                        # 如果满足条件，记录下当前unique_identification以及索引
                        if row.get('master_data') == 'YES' and float(row.get('amount', '0')) != 0:
                            unique_id_row_indices[unique_id] = index
                        else:
                            row['is_master_data'] = 0
                        # 检查之前是否有这个 unique_id 被标记，如果有，添加 'relate_uuid' 字段
                        if unique_id in unique_id_row_indices:
                            row['relate_uuid'] = unique_id
                        finance, finance_err = FinanceTransactService.add(row)
                        if finance_err:
                            failures.append({
                                'index': index,
                                'error': f"财务添加失败 {finance_err}",
                                'data': dict(row)
                            })
                            continue
                    elif module == "thread":
                        if not row.get("customer_code", ""):
                            continue
                        user_name, err = DetailInfoService.get_detail(
                            search_params={"full_name": row.get("full_name", "")})
                        if user_name:
                            user_id = user_name['user_id']
                        else:
                            # 如果不存在该客户创建
                            row['user_type'] = "UNKNOWN"
                            user, user_err = UserService.user_add(row)
                            if user_err:
                                failures.append({
                                    'index': index,
                                    'error': f"用户添加失败 {user_err}",
                                    'data': dict(row)
                                })
                            user_id = user['user_id']
                            # 创建成功绑定用户企业关系
                            user_group, user_group_err = UserGroupService.user_bind_groups(user_id, group_id)
                            if user_group_err:
                                failures.append({
                                    'index': index,
                                    'error': f"用户添加失败 {user_err}",
                                    'data': dict(row)
                                })
                        row['user_id'] = user_id
                        row['category_id'] = category_id
                        thread_detail, thread_err_detail = ThreadItemService.detail(
                            search_params={"customer_code": row.get("customer_code", ""), "category_id": category_id},
                            sort="id")
                        if thread_detail:
                            row['main_thread_id'] = thread_detail['id']
                            row['is_subitem_thread'] = 1
                        thread, thread_err = ThreadItemService.add(row)
                        if thread_err:
                            failures.append({
                                'index': index,
                                'error': f"合同添加失败 {thread_err}",
                                'data': dict(row)
                            })
                    elif module == "user":
                        if row.get("user_type", "") == "个人":
                            row['user_type'] = "PERSON"
                        elif row.get("user_type", "") == "公司":
                            row['user_type'] = " COMPANY"
                        else:
                            row['user_type'] = "UNKNOWN"

                        user_name, err = DetailInfoService.get_detail(
                            search_params={"full_name": row.get("full_name", "")})
                        if user_name:
                            continue
                        user, user_err = UserService.user_add(row)
                        if user_err:
                            failures.append({
                                'index': index,
                                'error': f"用户添加失败 {user_err}",
                                'data': dict(row)
                            })
                except exc.SQLAlchemyError as e:
                    failures.append({
                        'index': index,
                        'error': str(e),
                        'data': row
                    })

            # 每处理一定数量的行，更新任务的进度
            task = MigrateImportProgress.objects.get(task_id=task_id)
            task.progress = 100
            task.status = 'SUCCESS' if not failures else 'FAILURE'
            task.save()
            if failures:
                print("未能执行以下记录:")
                now = datetime.datetime.now()  # 获取当前日期和时间
                timestamp = now.strftime("%Y%m%d%H%M%S")  # 格式化日期和时间
                filepath = import_logs + f"error_log_{timestamp}.xlsx"  # 在文件名中添加时间戳

                # 检查文件是否已存在
                # 尝试打开文件，如果文件不存在，则设置append_data为False
                # 如果文件已存在，将数据附加到现有工作簿中
                if os.path.isfile(filepath):
                    book = load_workbook(filepath)
                    writer = pd.ExcelWriter(filepath, engine='openpyxl')
                    writer.book = book
                    append_data = True
                else:
                    # 如果文件不存在，创建新的Excel文件
                    writer = pd.ExcelWriter(filepath, engine='openpyxl')
                    append_data = False

                sheet_name = 'Sheet1'  # 我们要写入的Sheet的名称
                for failure in failures:
                    failure_count += 1
                    print(f"索引: {failure['index']}, 错误: {failure['error']}, 数据: {failure['data']}")
                    data = failure['data']
                    data['索引'] = failure['index']  # 将索引添加到数据字典中
                    data['错误'] = failure['error']  # 将错误消息添加到数据字典中

                    df = pd.DataFrame(data, index=[0])  # 将字典转换为DataFrame

                    startrow = writer.sheets[sheet_name].max_row if append_data else 0
                    df.to_excel(writer, index=False, header=not append_data, startrow=startrow, sheet_name=sheet_name)
                    append_data = True  # 设置标志以便将来的数据附加到工作簿中
                writer.save()  # 在所有数据都已写入后，保存工作簿
            return {
                'status': 'success',
                'total': total,
                'success_count': success_count,
                'failure_count': failure_count,
                # 'failures': failures,
            }, None

        except Exception as e:
            logger.error(f"Error in bind_data_processing: {str(e)}")
            logger.error(traceback.format_exc())
            return {
                'status': 'error',
                'message': str(e),
            }, str(e)

    @staticmethod
    def data_processing(execl_file, module, import_category=False):
        config_set = MigrateExeclConfig.objects.filter(module=module).first()
        if not config_set:
            return None, "配置不存在"
        config = model_to_dict(config_set)
        database_config = config.get("database_config", None)
        # 如果不存在 则读取配置文件中的数据（为保证安全，数据库里不应该存储未加密的数据库信息）
        if not database_config:
            database_config = {}
            database_config['user'] = user
            database_config['password'] = password
            database_config['host'] = host
            database_config['port'] = port
            database_config['database'] = database

        data_mappings = config.get("data_mappings", None)
        if not data_mappings:
            return None, "数据映射不能为空"
        if isinstance(config.get("date_fields", ""), str):
            if config.get("date_fields", "").strip() == "":
                date_fields = []
            elif "," in config.get("date_fields", ""):
                date_fields = [str(num) for num in config.get("date_fields", "").split(",")]
            elif "，" in config.get("date_fields", ""):
                date_fields = [str(num) for num in config.get("date_fields", "").split("，")]
            else:
                date_fields = [str(config.get("date_fields", ""))]
        elif isinstance(config.get("date_fields", ""), int):
            date_fields = [str(config.get("date_fields", ""))]
        else:
            date_fields = []

        if module == "finance":
            data_mappings[0]['generated_fields']["subtitle"] = unique_identification
        execl_import_set = execl_import(database_config, execl_file, date_fields, data_mappings, import_logs)
        return execl_import_set, None
