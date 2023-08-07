from main.settings import BASE_DIR
from pathlib import Path
from django.forms.models import model_to_dict
from sqlalchemy import create_engine, Table, MetaData, exc, insert, text, update
import pandas as pd
import openpyxl
import datetime
from openpyxl.utils import get_column_letter
from ..models import MigrateExeclConfig
from ..utils.j_config import JConfig
from ..utils.j_dict import JDict

module_root = str(Path(__file__).resolve().parent)
# 配置之对象
main_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="main"))
module_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="main"))

import_main_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="xj_migrate"))
import_module_config_dict = JDict(JConfig.get_section(path=str(BASE_DIR) + "/config.ini", section="xj_migrate"))

# 获取数据库信息
user = main_config_dict.mysql_user or module_config_dict.mysql_user or ""
password = main_config_dict.mysql_password or module_config_dict.mysql_password or ""
host = main_config_dict.mysql_host or module_config_dict.mysql_host or ""
port = main_config_dict.mysql_port or module_config_dict.mysql_port or ""
database = main_config_dict.mysql_database or module_config_dict.mysql_database or ""

import_logs = import_main_config_dict.import_logs or import_module_config_dict.import_logs or ""

"""
       execl表格导入，支持合并单元格，支持单表和多表的导入。
       PS : execl数据导入时，需要进行data_mappings的配置
            多表导入不支持扩展字段导入
       @:param database_config 数据库配置
                database_config = {
                    "user": "用户名",
                    "password": "密码",
                    "host": "地址",
                    "port": "端口号",
                    "database": "数据库"
                }
       @:param execl_file execl文件路径
       @:param data_mappings 数据配置，结构如下：
            data_mappings = [
                {
                    "table_name": "thread", #导入的数据库表
                    "fields": {
                        "工程项目": "title",    #key的值对应execl表头的值，value对应数据库表的字段
                        "合同编码": "subtitle"
                    },
                    #定义默认生成规则 #key对应数据库表的字段,value对应规则方法
                     "generated_fields": {
                        "subtitle": generate_code,
                        "author": generate
                    },
                    "foreign_keys": {}
                },
                {
                    "table_name": "invoice_invoice",
                    "fields": {
                        "开票时间": "invoice_time",
                        "发票号码": "invoice_number",
                        "开票金额": "invoice_price",
                        "税金点数": "tax_rate",
                    },
                    #如果是多表导入的情况，会先插入thread的数据，然后返回新生成的主键id，作为外键插入invoice_invoice表
                    "foreign_keys": {
                        "thread_id": "thread",
                    },
                    #配置这里，会到invoice_type表里根据execl里表头为 “专/普”里的值去和 “description”字段的值进行匹配，成功后返回，主键id，作为外键插入invoice_invoice表
                    "unique_foreign_keys": {
                        "invoice_type_id": { #外键名
                            "related_table": "invoice_type",#对应的表名
                            "unique_field_excel": "专/普",#execl表头名
                            "unique_field_db": "description" #数据库表对应字段名
                        }
                    }
                },
                {
                    "table_name": "finance_transact",
                    "fields": {
                        "摘要": "summary",
                        "汇入时间": "transact_time"
                    },
                    #扩展表主键id 用来关联主表和扩展表的映射关系 扩展表写入必传参数
                    "extend_foreign_keys": "finance_id",
                    #扩展表映射配置
                    "extend_fields": {
                        "import_amount": { #扩展字段名
                            "related_table": "finance_extend_field",#扩展表
                            "mapping_table": "finance_extend_data",#扩展映射数据表
                            "unique_field_excel": "汇入金额" #execl表头名
                        },
                        "export_price": {
                            "related_table": "finance_extend_field",
                            "mapping_table": "finance_extend_data",
                            "unique_field_excel": "汇出金额"
                        }
                    },
                    "unique_foreign_keys": {  
                        "thread_id": {
                            "related_table": "thread",
                            "unique_field_excel": "合同编码",
                            "unique_field_db": "subtitle"
                        },
                    },
                    "accumulate_field": "是否累加" #execl表头名 用来判断是否进行累加操作
            ]
       @:param date_fields 需要处理的时间字段
       @:param import_category True: 多表导入, False: 单表导入 默认单表

"""


class ExeclService:
    @staticmethod
    def data_processing(execl_file, module, import_category=False):
        config_set = MigrateExeclConfig.objects.filter(module=module).first()
        if not config_set:
            return None, "配置不存在"
        config = model_to_dict(config_set)
        database_config = config.get("database_config", None)
        # 如果不存在 则读取配置文件中的数据（为保证安全，数据库里不应该存储未加密的数据库信息）
        if not database_config:
            database_config = {}
            database_config['user'] = user
            database_config['password'] = password
            database_config['host'] = host
            database_config['port'] = port
            database_config['database'] = database

        data_mappings = config.get("data_mappings", None)
        if not data_mappings:
            return None, "数据映射不能为空"

        if isinstance(config.get("date_fields", None), str):
            if config.get("date_fields", None).strip() == "":
                date_fields = []
            elif "," in config.get("date_fields", None):
                date_fields = [str(num) for num in config.get("date_fields", None).split(",")]
            elif "，" in config.get("date_fields", None):
                date_fields = [str(num) for num in config.get("date_fields", None).split("，")]
            else:
                date_fields = [str(config.get("date_fields", None))]
        elif isinstance(config.get("date_fields", None), int):
            date_fields = [str(config.get("date_fields", None))]
        else:
            return None, "不支持的值类型,应为字符串或int"

        # 数据库连接信息
        DATABASE_URI = "mysql+mysqldb://" + database_config['user'] + ":" + database_config[
            'password'] + "@" + database_config['host'] + ":" + database_config[
                           'port'] + "/" + database_config[
                           'database'] + "?charset=utf8"  # 替换为你的数据库URI

        engine = create_engine(DATABASE_URI, echo=True)  # 设置echo=True以打印所有执行的SQL语句

        metadata = MetaData()

        # Excel文件路径
        wb = openpyxl.load_workbook(execl_file)
        sheet = wb.active

        # 复制合并单元格数据
        merged_cells_ranges = list(sheet.merged_cells.ranges)

        # 遍历所有合并单元格
        for merged_cell in merged_cells_ranges:
            # 获取合并单元格的起始行、列和结束行、列
            min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col

            # 获取合并单元格的数据
            value = sheet.cell(row=min_row, column=min_col).value

            # 将行列索引转换为单元格引用，例如"A1:B2"
            range_string = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

            # 拆分合并单元格
            sheet.unmerge_cells(range_string)

            # 将数据重新填充到原始合并的单元格中
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = value

            # 提取数据并以列表形式输出
            data = []
            for row in sheet.iter_rows(values_only=True):
                formatted_row = []
                for cell in row:
                    if isinstance(cell, (datetime.datetime, datetime.date)):
                        formatted_row.append(cell.strftime("%Y-%m-%d"))
                    else:
                        formatted_row.append(cell)
                data.append(formatted_row)

        data = list(sheet.values)
        # 创建 DataFrame，并将所有列的类型转为字符串
        df = pd.DataFrame(data[1:], columns=data[0])

        # 数据预处理
        df = df.astype(str)
        df.replace('nan', '', inplace=True)  # 用空字符串替换NaN
        for col in date_fields:
            df[col] = df[col].replace('', '0000-00-00')  # 使用特定日期替换空字符串

        filepath = ""
        success_count = 0
        failure_count = 0
        failures = []
        foreign_keys = {}
        mapping_table = None

        # 在数据预处理后的位置
        if not import_category:
            # 数据预处理后
            for index, row in df.iterrows():
                for mapping in data_mappings:
                    table_name = mapping['table_name']
                    fields = mapping['fields']
                    generated_fields = mapping.get('generated_fields', {})
                    table = Table(table_name, metadata, autoload_with=engine)

                    unique_foreign_key_present = True
                    foreign_keys = {}
                    for unique_fk, fk_conf in mapping.get('unique_foreign_keys', {}).items():
                        related_table = Table(fk_conf['related_table'], metadata, autoload_with=engine)
                        unique_field_excel = fk_conf['unique_field_excel']
                        unique_field_db = fk_conf['unique_field_db']
                        with engine.begin() as connection:
                            result = connection.execute(
                                text(f"SELECT id FROM {related_table.name} WHERE {unique_field_db} = :value"),
                                {"value": row[unique_field_excel]}
                            )
                            fk_id = result.scalar()
                            if fk_id is None:
                                unique_foreign_key_present = False
                                break
                            else:
                                foreign_keys[unique_fk] = fk_id

                    if not unique_foreign_key_present:
                        # 如果不存在唯一外键，则跳过此记录 并将其添加到故障列表中
                        failures.append({
                            'index': index,
                            'error': f"唯一外键 {unique_fk} 未找到值 {row[unique_field_excel]}",
                            'data': dict(row)
                        })
                        continue
                    record = {}
                    for cn, field in fields.items():
                        if cn in row:
                            if cn in date_fields:
                                date_value = pd.to_datetime(row[cn], errors='coerce')
                                if pd.isna(date_value):  # 处理缺少的日期值
                                    record[field] = None
                                else:
                                    record[field] = date_value
                            else:
                                record[field] = row[cn]

                    for field, generator in generated_fields.items():
                        record[field] = generator()

                    # 处理 foreign_keys 和 unique_foreign_keys
                    for fk_field, fk_table in mapping.get('foreign_keys', {}).items():
                        if fk_table in foreign_keys:
                            record[fk_field] = foreign_keys[fk_table]

                    for fk_field, _ in mapping.get('unique_foreign_keys', {}).items():
                        if fk_field in foreign_keys:
                            record[fk_field] = foreign_keys[fk_field]
                    try:
                        with engine.begin() as connection:
                            # 检查记录是否存在，累积字段为“是” 则累加否则覆盖
                            if 'unique_fk' in locals() or 'unique_fk' in globals():
                                if (mapping.get("accumulate_field", None) not in row or row[
                                    mapping["accumulate_field"]] != "是") and \
                                        foreign_keys[unique_fk] is not None:
                                    result = connection.execute(
                                        text(f"SELECT * FROM {table.name} WHERE {unique_fk} = :value"),
                                        {"value": foreign_keys[unique_fk]}
                                    )
                                    record_exists = result.fetchone() is not None
                                else:
                                    record_exists = False
                            else:
                                record_exists = False

                            if record_exists:
                                # 记录存在，执行更新操作
                                stmt = (
                                    update(table)
                                    .where(table.c[unique_fk] == foreign_keys[unique_fk])
                                    .values(record)
                                )
                                connection.execute(stmt)
                                result = connection.execute(
                                    text(f"SELECT id FROM {table.name} WHERE {unique_fk} = :value"),
                                    {"value": foreign_keys[unique_fk]}
                                )
                                row = result.fetchone()
                                if row is not None:
                                    foreign_keys[mapping.get('extend_foreign_keys')] = row[0]
                            else:
                                # 不存在，执行插入操作
                                result = connection.execute(insert(table).values(record))
                                if len(table.primary_key.columns) == 1:
                                    column = list(table.primary_key.columns)[0]
                                    if column.autoincrement:
                                        result = connection.execute(text('SELECT LAST_INSERT_ID()'))
                                        last_id = result.scalar()
                                        if last_id is not None:
                                            foreign_keys[mapping.get('extend_foreign_keys')] = last_id
                            if mapping.get('extend_foreign_keys') in foreign_keys:
                                data_to_insert = {
                                    mapping.get('extend_foreign_keys', ""): foreign_keys.get(
                                        mapping.get('extend_foreign_keys')),
                                }
                                for extend_field, conf in mapping.get('extend_fields', {}).items():
                                    related_table = Table(conf['related_table'], metadata, autoload_with=engine)
                                    mapping_table = Table(conf['mapping_table'], metadata, autoload_with=engine)
                                    unique_field_excel = conf['unique_field_excel']
                                    result = connection.execute(
                                        text(f"SELECT field_index FROM {related_table.name} WHERE field = :value"),
                                        {"value": extend_field}
                                    )
                                    result_row = result.fetchone()
                                    if result_row is not None:
                                        field_index = result_row[0]
                                        # 更新data_to_insert字典
                                        data_to_insert[field_index] = df.at[index, unique_field_excel]

                            # 循环后
                            if mapping_table is not None:
                                # 检查foreign_keys字典中是否存在外键
                                extend_foreign_key = mapping.get('extend_foreign_keys')
                                if extend_foreign_key not in foreign_keys or foreign_keys[extend_foreign_key] is None:
                                    print(
                                        f"在Foreign_keys中找不到外键 {extend_foreign_key} 跳过行 {index}")
                                    continue
                                # 检查记录是否存在
                                result = connection.execute(
                                    text(f"SELECT * FROM {mapping_table.name} WHERE {extend_foreign_key} = :value"),
                                    {"value": foreign_keys[extend_foreign_key]}
                                )
                                record_exists = result.fetchone() is not None
                                if record_exists:
                                    data_to_insert.pop(mapping.get('extend_foreign_keys', ""))
                                    # 记录存在，执行更新操作
                                    stmt = (
                                        update(mapping_table)
                                        .where(mapping_table.c[extend_foreign_key] == foreign_keys[extend_foreign_key])
                                        .values(data_to_insert)
                                    )
                                    connection.execute(stmt)
                                else:
                                    # 不存在，执行插入操作
                                    connection.execute(insert(mapping_table).values(data_to_insert))

                        success_count += 1
                    except exc.SQLAlchemyError as e:
                        failures.append({
                            'index': index,
                            'error': str(e),
                            'data': record
                        })
        else:
            for index, row in df.iterrows():
                for mapping in data_mappings:
                    table_name = mapping['table_name']
                    fields = mapping['fields']
                    generated_fields = mapping.get('generated_fields', {})

                    if 'foreign_keys' not in mapping and 'unique_foreign_keys' not in mapping:  # 处理没有外键的表
                        continue

                    table = Table(table_name, metadata, autoload_with=engine)
                    record = {}
                    for cn, field in fields.items():
                        if cn in row:
                            if cn in date_fields:
                                date_value = pd.to_datetime(row[cn], errors='coerce')
                                if pd.isna(date_value):  # 处理缺少的日期值
                                    record[field] = None
                                else:
                                    record[field] = date_value
                            else:
                                record[field] = row[cn]

                    for field, generator in generated_fields.items():
                        record[field] = generator()

                    for fk_field, fk_info in mapping.get('unique_foreign_keys', {}).items():
                        related_table = fk_info["related_table"]
                        unique_field_excel = fk_info["unique_field_excel"]
                        unique_field_db = fk_info["unique_field_db"]
                        related_table = Table(related_table, metadata, autoload_with=engine)

                        with engine.begin() as connection:
                            result = connection.execute(
                                text(f"SELECT id FROM {related_table.name} WHERE {unique_field_db} = :value"),
                                {"value": row[unique_field_excel]}
                            )
                            id_ = result.scalar()
                            record[fk_field] = id_

                    for fk_field, fk_table in mapping.get('foreign_keys', {}).items():
                        if fk_table in foreign_keys:
                            record[fk_field] = foreign_keys[fk_table]

                    try:
                        with engine.begin() as connection:
                            result = connection.execute(insert(table).values(record))
                            if len(table.primary_key.columns) == 1:
                                column = list(table.primary_key.columns)[0]
                                if column.autoincrement:
                                    result = connection.execute(text('SELECT LAST_INSERT_ID()'))
                                    last_id = result.scalar()
                                    foreign_keys[table_name] = last_id
                        success_count += 1
                    except exc.SQLAlchemyError as e:
                        failures.append({
                            'index': index,
                            'error': str(e),
                            'data': record
                        })
        if failures:
            print("未能执行以下记录:")
            now = datetime.datetime.now()  # 获取当前日期和时间
            timestamp = now.strftime("%Y%m%d%H%M%S")  # 格式化日期和时间
            filepath = import_logs + f"error_log_{timestamp}.txt"  # 在文件名中添加时间戳
            with open(filepath, "w", encoding='utf-8') as f:  # 使用 UTF-8 编码打开文件
                for failure in failures:
                    failure_count += 1
                    print(f"索引: {failure['index']}, 错误: {failure['error']}, 数据: {failure['data']}")
                    f.write(f"索引: {failure['index']}, 错误: {failure['error']}, 数据: {failure['data']}\n")
            print(f"错误数据已保存到文件：{filepath}")

        if filepath:
            return {"message": "已成功执行" + str(success_count) + "条记录，失败" + str(failure_count) + "条",
                    "error_path": filepath}, None
        return {"message": "已成功执行" + str(success_count) + "条记录，失败" + str(failure_count) + "条"}, None

