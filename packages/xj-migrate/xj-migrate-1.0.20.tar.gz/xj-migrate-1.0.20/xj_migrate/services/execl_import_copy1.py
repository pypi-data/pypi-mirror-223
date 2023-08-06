from sqlalchemy import create_engine, MetaData
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import datetime

from xj_migrate.utils.utility_method import replace_dict_key, replace_key_in_dict, replace_key_in_list_dicts, \
    aggregate_data, replace_key_in_list_replacement_dicts

# 预先定义的数据映射

# 预先定义的数据映射
# data_mappings = {
#     "客户新编码": {"db_field": "order_no", "type": "remain"},
#     "开票时间": {"db_field": "invoice_time", "type": "remain"},
#     "发票号码": {"db_field": "invoice_number", "type": "remain"},
#     "开票金额": {"db_field": "invoice_price", "type": "remain"},
#     # "税率": {"db_field": "tax_rate", "type": "extract", "additional_value": {}},
#     # "检索发票税额": {"db_field": "invoice_tax", "type": "extract", "additional_value": {"sand_box": "bbb"}},
#     "税率": {"db_field": "tax_rate", "type": "remain"},
#     "检索发票税额": {"db_field": "invoice_tax", "type": "remain"},
#     "专/普": {"db_field": "invoice_type", "type": "remain"}
# }
# old_keys = ['invoice_price']
# new_key = 'amount'


# old_keys = ['remittance_amount', 'management_fees', 'taxes', 'commission', 'amount_remitted']
# new_key = 'amount'

replacement_dict = {
    'remittance_amount': "amount",
    'management_fees': "amount",
    'taxes': "amount",
    'commission': "amount",
    'amount_remitted': "amount",
    'deduction_basis': "amount",
    "remit_time": "transact_time"
}

data_mappings = {
    "客户新编码": {"db_field": "order_no", "type": "remain"},
    "汇入时间": {"db_field": "transact_time", "type": "remain"},
    "汇入金额": {"db_field": "remittance_amount", "type": "remain"},
    "扣费依据": {"db_field": "deduction_basis", "type": "extract",
                 "additional_value": {"sand_box": "DEDUCTION_BASIS", "order_no": ""}},
    "管理费点数": {"db_field": "manage_point", "type": "remain"},
    "管理费": {"db_field": "management_fees", "type": "extract",
               "additional_value": {"order_no": "", "manage_point": "",
                                    "sand_box": "MANAGEMENT_FEE_RECEIVABLE"}},
    "税金点数": {"db_field": "tax_point", "type": "remain"},
    "税金": {"db_field": "taxes", "type": "extract",
             "additional_value": {"sand_box": "TAX_RECEIVABLES", "order_no": "", "tax_point": ""}},
    "佣金点数": {"db_field": "brokerage_point", "type": "remain"},
    "佣金": {"db_field": "commission", "type": "extract",
             "additional_value": {"order_no": "", "sand_box": "COMMISSION_RECEIVABLE",
                                  "brokerage_point": ""}},
    "摘要": {"db_field": "summary", "type": "remain"},
    "汇出时间": {"db_field": "remit_time", "type": "remain"},
    "汇出金额": {"db_field": "amount_remitted", "type": "extract",
                 "additional_value": {"order_no": "", "remit_time": ""},
                 "negate": True},
}
# 读取 Excel 文件
wb = openpyxl.load_workbook('C:/Users/Lenovo/Desktop/4.xlsx')
sheet = wb.active

# 拆分合并单元格并提取数据
merged_cells_ranges = list(sheet.merged_cells.ranges)
for merged_cell in merged_cells_ranges:
    min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
    value = sheet.cell(row=min_row, column=min_col).value
    range_string = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    sheet.unmerge_cells(range_string)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            sheet.cell(row=row, column=col).value = value

# 从 Excel 中获取数据并创建 DataFrame
data = list(sheet.values)
df = pd.DataFrame(data[1:], columns=data[0])

# 数据预处理
df = df.astype(str)
df.replace('nan', '', inplace=True)

# 对 DataFrame 的每一行应用数据映射，生成对应的数据库插入数据
db_data = []

# fields_to_extract = ["税率", "检索发票税额"]  # 需要拆分出来的字段
# fields_remain = [k for k in data_mappings.keys() if k not in fields_to_extract]  # 保持在原来字典的字段
# additional_values = {"税率": {"sand_box": "aaa"}, "检索发票税额": {"sand_box": "bbb"}}  # 额外的字段和值
# for _, row in df.iterrows():
#     row_data_all = {data_mappings[k]: v for k, v in row.to_dict().items() if k in data_mappings}
#
#     # 先添加包含部分字段的字典
#     db_data.append({data_mappings[k]: row_data_all[data_mappings[k]] for k in fields_remain})
#
#     # 再分别为需要拆分的字段创建新的字典并添加到列表中
#     for k in fields_to_extract:
#         db_data.append({**{data_mappings[k]: row_data_all[data_mappings[k]]}, **additional_values[k]})
for _, row in df.iterrows():
    row_data_remain = {}
    extracted_data = []
    remain_fields_values = {}  # 保存所有 remain 类型字段的值
    for k, v in row.to_dict().items():
        if k in data_mappings:
            if v != 'None':
                # 如果字段被标记为需要取负，那么就添加负号
                if "negate" in data_mappings[k] and data_mappings[k]["negate"]:
                    try:
                        v = str(-float(v))  # 尝试转换为浮点数并取负
                    except ValueError:
                        pass  # 如果转换失败，保持原样

                if data_mappings[k]["type"] == "remain":
                    row_data_remain[data_mappings[k]["db_field"]] = v
                    remain_fields_values[data_mappings[k]["db_field"]] = v
                elif data_mappings[k]["type"] == "extract":
                    additional_value = data_mappings[k]["additional_value"].copy()  # 创建additional_value的副本以避免修改原始字典
                    # 填充 additional_value
                    for field in additional_value.keys():
                        if field in remain_fields_values:
                            additional_value[field] = remain_fields_values[field]
                    extracted_data.append({**{data_mappings[k]["db_field"]: v}, **additional_value})
    db_data.append(row_data_remain)
    db_data.extend(extracted_data)

# print(aggregate_data(db_data, "order_no", ['invoice_tax', 'invoice_price']))
# db_data 就是你要的列表，你可以用它来插入数据到数据库
db_data = replace_key_in_list_replacement_dicts(db_data, replacement_dict)

for item in db_data:
    print(item)
