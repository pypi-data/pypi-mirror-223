# Copyright (c) 2022-2023 IO-Aero. All rights reserved. Use of this
# source code is governed by the IO-Aero License, that can
# be found in the LICENSE.md file.

"""Managing the database schema of the PostgreSQL database."""
import json
import os
import platform
import shutil
import subprocess
import zipfile
from datetime import datetime
from datetime import timezone
from pathlib import Path

import pyodbc  # type: ignore
import requests
from openpyxl.reader.excel import load_workbook
from psycopg2.errors import ForeignKeyViolation  # pylint: disable=no-name-in-module
from psycopg2.errors import UniqueViolation  # pylint: disable=no-name-in-module
from psycopg2.extensions import connection
from psycopg2.extensions import cursor

from ioavstatsdb import db_utils
from ioavstatsdb import io_config
from ioavstatsdb import io_glob
from ioavstatsdb import io_utils

IO_LAST_SEEN = datetime.now(timezone.utc)


# ------------------------------------------------------------------
# Check for DDL changes.
# ------------------------------------------------------------------
# flake8: noqa
# pylint: disable=too-many-lines
def _check_ddl_changes(msaccess: str) -> None:
    """Check for DDL changes."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    download_work_dir = os.path.join(
        os.getcwd(), io_config.settings.download_work_dir.replace("/", os.sep)
    )

    shutil.copy(
        os.path.join(
            download_work_dir.replace("/", os.sep),
            msaccess + "." + io_glob.FILE_EXTENSION_MDB,
        ),
        os.path.join(
            download_work_dir.replace("/", os.sep),
            io_config.settings.razorsql_profile + "." + io_glob.FILE_EXTENSION_MDB,
        ),
    )

    msaccess_mdb = os.path.join(
        download_work_dir.replace("/", os.sep),
        "IO-AVSTATS.mdb",
    )

    # INFO.00.051 msaccess_file='{msaccess_file}'
    io_utils.progress_msg(io_glob.INFO_00_051.replace("{msaccess_file}", msaccess_mdb))

    if not os.path.exists(msaccess_mdb):
        # ERROR.00.932 File '{filename}' is not existing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_932.replace("{filename}", msaccess_mdb)
        )

    msaccess_sql = os.path.join(
        download_work_dir.replace("/", os.sep),
        msaccess + "." + io_glob.FILE_EXTENSION_SQL,
    )

    # INFO.00.051 msaccess_file='{msaccess_file}'
    io_utils.progress_msg(io_glob.INFO_00_051.replace("{msaccess_file}", msaccess_sql))

    razorsql_jar_file = ""
    razorsql_java_path = ""

    if platform.system() == "Windows":
        razorsql_jar_file = io_config.settings.razorsql_jar_file_windows
        razorsql_java_path = io_config.settings.razorsql_java_path_windows
    elif platform.system() == "Linux":
        razorsql_jar_file = io_config.settings.razorsql_jar_file_linux
        razorsql_java_path = io_config.settings.razorsql_java_path_linux
    else:
        # ERROR.00.908 The operating system '{os}' is not supported
        io_utils.terminate_fatal(
            io_glob.ERROR_00_908.replace("{os}", platform.system())
        )

    # INFO.00.052 razorsql_jar_file='{razorsql_jar_file}'
    io_utils.progress_msg(
        io_glob.INFO_00_052.replace("{razorsql_jar_file}", razorsql_jar_file)
    )

    if not Path(razorsql_jar_file).is_file():
        # ERROR.00.932 File '{filename}' is not existing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_932.replace("{filename}", razorsql_jar_file)
        )

    # INFO.00.053 razorsql_java_path='{razorsql_java_path}'"
    io_utils.progress_msg(
        io_glob.INFO_00_053.replace("{razorsql_java_path}", razorsql_java_path)
    )

    subprocess.run(
        [
            razorsql_java_path,
            "-jar",
            razorsql_jar_file,
            "-backup",
            io_config.settings.razorsql_profile,
            "null",
            "null",
            ";",
            "null",
            msaccess_sql,
            "NO",
            "tables",
            "YES",
            "null",
            "NO",
            "NO",
        ],
        check=True,
    )

    # INFO.00.011 The DDL script for the MS Access database '{msaccess}'
    # was created successfully
    io_utils.progress_msg(io_glob.INFO_00_011.replace("{msaccess}", msaccess))

    if msaccess == io_glob.MSACCESS_PRE2008:
        # INFO.00.020 The DDL script for the MS Access database '{msaccess}'
        # must be checked manually
        io_utils.progress_msg(io_glob.INFO_00_020.replace("{msaccess}", msaccess))
    else:
        _compare_ddl(msaccess)

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Compare the schema definitions with the reference file.
# ------------------------------------------------------------------
def _compare_ddl(msaccess: str) -> None:
    """Compare the schema definitions wirth the reference file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    reference_filename = os.path.join(
        io_config.settings.razorsql_reference_dir.replace("/", os.sep),
        io_config.settings.razorsql_reference_file,
    )

    # pylint: disable=consider-using-with
    reference_file = open(
        reference_filename, "r", encoding=io_glob.FILE_ENCODING_DEFAULT
    )
    reference_file_lines = reference_file.readlines()
    reference_file.close()

    msaccess_filename = os.path.join(
        io_config.settings.download_work_dir.replace("/", os.sep),
        msaccess + "." + io_glob.FILE_EXTENSION_SQL,
    )

    # pylint: disable=consider-using-with
    msaccess_file = open(msaccess_filename, "r", encoding=io_glob.FILE_ENCODING_DEFAULT)
    msaccess_file_lines = msaccess_file.readlines()
    msaccess_file.close()

    if len(reference_file_lines) != len(msaccess_file_lines):
        # ERROR.00.911 Number of lines differs: file '{filename}' lines
        # {filename_lines} versus file'{reference}' lines {reference_lines}
        io_utils.terminate_fatal(
            io_glob.ERROR_00_911.replace("{filename}", msaccess_filename)
            .replace("{filename_lines}", str(len(msaccess_file_lines)))
            .replace("{reference}", reference_filename)
            .replace("{reference_lines}", str(len(reference_file_lines)))
        )

    for line_no, line in enumerate(reference_file_lines):
        if line != msaccess_file_lines[line_no]:
            # INFO.00.009 line no.: {line_no}
            # INFO.00.010 {status} '{line}'
            io_utils.progress_msg_core(
                io_glob.INFO_00_009.replace("{line_no}", str(line_no))
            )
            io_utils.progress_msg_core(
                io_glob.INFO_00_010.replace("{status}", "expected").replace(
                    "{line}", line
                )
            )
            io_utils.progress_msg_core(
                io_glob.INFO_00_010.replace("{status}", "received").replace(
                    "{line}", msaccess_file_lines[line_no]
                )
            )
            # ERROR.00.910 The schema definition in file'{filename}'
            # does not match the reference definition in file'{reference}'
            io_utils.terminate_fatal(
                io_glob.ERROR_00_910.replace("{filename}", msaccess_filename).replace(
                    "{reference}", reference_filename
                )
            )

    # INFO.00.012 The DDL script for the MS Access database '{msaccess}'
    # is identical to the reference script
    io_utils.progress_msg(io_glob.INFO_00_012.replace("{msaccess}", msaccess))

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Delete all the NTSB data.
# ------------------------------------------------------------------
def _delete_ntsb_data(
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Delete all the NTSB data."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    for table in [
        # Level 4 - FK: ev_id & Aircraft_Key & crew_no
        "dt_flight_crew",
        "flight_time",
        # Level 4 - FK: ev_id & Aircraft_Key & Occurrence_No
        "seq_of_events",
        # Level 3 - FK: ev_id & Aircraft_Key
        "dt_aircraft",
        "engines",
        "events_sequence",
        "findings",
        "flight_crew",
        "injury",
        "narratives",
        "occurrences",
        # Level 2 - FK: ev_id
        "aircraft",
        "dt_events",
        "ntsb_admin",
        # Level 1 - without FK
        "events",
    ]:
        cur_pg.execute(
            f"""
        SELECT count(*)
          FROM {table}
            """,
        )
        row_pg = cur_pg.fetchone()
        no_rows = row_pg["count"] if row_pg else 0  # type: ignore
        if no_rows > 0:
            cur_pg.execute(
                f"""
            TRUNCATE {table} CASCADE;
            """,
            )
            io_utils.progress_msg(
                f"Table {table:<15} - number rows deleted : {str(no_rows):>8}"
            )
            conn_pg.commit()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load airport data from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def _load_airport_data() -> None:
    io_glob.logger.debug(io_glob.LOGGER_START)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Delete existing data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_airports
    """,
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()
        io_utils.progress_msg("-" * 80)
        # INFO.00.087 Database table io_airports: Delete the existing data
        io_utils.progress_msg(io_glob.INFO_00_087)
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")

    # ------------------------------------------------------------------
    # Start processing airport data.
    # ------------------------------------------------------------------

    us_states: list[str] = _sql_query_us_states(conn_pg)

    locids: list[str] = _load_npias_data(us_states)

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_faa_airports_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.943 The airport file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_943.replace("{filename}", filename_excel)
        )

    # INFO.00.089 Database table io_airports: Load data from file '{filename}'
    io_utils.progress_msg("-" * 80)
    io_utils.progress_msg(io_glob.INFO_00_089.replace("{filename}", filename_excel))

    count_select = 0
    count_upsert = 0
    count_usable = 0

    x_idx = 0
    y_idx = 1
    global_idx = 3
    ident_idx = 4
    name_idx = 5
    latitude_idx = 6
    longitude_idx = 7
    elevation_idx = 8
    type_code_idx = 10
    servcity_idx = 11
    state_idx = 12
    country_idx = 13
    operstatus_idx = 14
    privateuse_idx = 15
    iapexists_idx = 16
    dodhiflib_idx = 17
    far91_idx = 18
    far93_idx = 19
    mil_code_idx = 20
    airanal_idx = 21

    # ------------------------------------------------------------------
    # Load the airport data.
    # ------------------------------------------------------------------

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        count_select += 1

        ident = row[ident_idx].value
        if ident is None or not ident in locids:
            continue

        country = row[country_idx].value
        if country != "UNITED STATES":
            continue

        dec_longitude = row[x_idx].value
        if dec_longitude == "X":
            continue

        state = row[state_idx].value
        if state is None or not state in us_states:
            continue

        count_usable += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        airanal = row[airanal_idx].value
        country = "USA"
        dec_latitude = row[y_idx].value
        dodhiflib = row[dodhiflib_idx].value
        elevation = row[elevation_idx].value
        far91 = row[far91_idx].value
        far93 = row[far93_idx].value
        global_id = row[global_idx].value
        iapexists = row[iapexists_idx].value
        latitude = row[latitude_idx].value
        longitude = row[longitude_idx].value
        mil_code = row[mil_code_idx].value
        name = row[name_idx].value
        operstatus = row[operstatus_idx].value
        privateuse = row[privateuse_idx].value
        servcity = row[servcity_idx].value
        state = row[state_idx].value
        type_code = row[type_code_idx].value

        cur_pg.execute(
            """
        INSERT INTO io_airports AS ia (
               global_id,
               airanal,
               country,
               dec_latitude,
               dec_longitude,
               dodhiflib,
               elevation,
               far91,
               far93,
               iapexists,
               ident,
               latitude,
               longitude,
               mil_code,
               name,
               operstatus,
               privateuse,
               servcity,
               state,
               type_code,
               first_processed
               ) VALUES (
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s,%s,
               %s
               )
        ON CONFLICT ON CONSTRAINT io_airports_pkey
        DO UPDATE
           SET airanal = %s,
               country = %s,
               dec_latitude = %s,
               dec_longitude = %s,
               dodhiflib = %s,
               elevation = %s,
               far91 = %s,
               far93 = %s,
               iapexists = %s,
               ident = %s,
               latitude = %s,
               longitude = %s,
               mil_code = %s,
               name = %s,
               operstatus = %s,
               privateuse = %s,
               servcity = %s,
               state = %s,
               type_code = %s,
               last_processed = %s
         WHERE ia.global_id = %s
        """,
            (
                global_id,
                airanal,
                country,
                dec_latitude,
                dec_longitude,
                dodhiflib,
                elevation,
                far91,
                far93,
                iapexists,
                ident,
                latitude,
                longitude,
                mil_code,
                name,
                operstatus,
                privateuse,
                servcity,
                state,
                type_code,
                datetime.now(),
                airanal,
                country,
                dec_latitude,
                dec_longitude,
                dodhiflib,
                elevation,
                far91,
                far93,
                iapexists,
                ident,
                latitude,
                longitude,
                mil_code,
                name,
                operstatus,
                privateuse,
                servcity,
                state,
                type_code,
                datetime.now(),
                global_id,
            ),
        )
        count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_sequence_of_events_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load aviation occurrence categories from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def _load_aviation_occurrence_categories() -> None:
    """Load aviation occurrence categories from an MS Excel file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_aviation_occurrence_categories_xlsx.replace(
            "/", os.sep
        ),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.937 The aviation occurrence categories file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_937.replace("{filename}", filename_excel)
        )

    # INFO.00.074 Database table io_aviation_occurrence_categories: Load data from file '{filename}'
    io_utils.progress_msg(
        io_glob.INFO_00_074.replace(
            "{filename}",
            io_config.settings.download_file_aviation_occurrence_categories_xlsx,
        )
    )
    io_utils.progress_msg("-" * 80)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    count_delete = 0
    count_upsert = 0
    count_select = 0

    # ------------------------------------------------------------------
    # Load the aviation occurrence categories.
    # ------------------------------------------------------------------

    cictt_code_idx = 0
    identifier_idx = 1
    definition_idx = 2

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        cictt_code = row[cictt_code_idx].value.upper().rstrip()
        if cictt_code == "CICTT CODE":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        identifier = row[identifier_idx].value.upper().rstrip()
        definition = row[definition_idx].value.rstrip()

        cur_pg.execute(
            """
        INSERT INTO io_aviation_occurrence_categories AS aoc (
               cictt_code,
               identifier,
               definition,
               first_processed,
               last_seen
               ) VALUES (
               %s,%s,%s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_aviation_occurrence_categories_pkey
        DO UPDATE
           SET identifier = %s,
               definition = %s,
               last_processed = %s,
               last_seen = %s
         WHERE aoc.cictt_code = %s
        """,
            (
                cictt_code,
                identifier,
                definition,
                datetime.now(),
                IO_LAST_SEEN,
                identifier,
                definition,
                datetime.now(),
                IO_LAST_SEEN,
                cictt_code,
            ),
        )
        count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    # ------------------------------------------------------------------
    # Delete the obsolete data.
    # ------------------------------------------------------------------

    count_select = 0

    # pylint: disable=line-too-long
    cur_pg.execute(
        """
    SELECT cictt_code
      FROM io_aviation_occurrence_categories
     WHERE last_seen <> %s
        """,
        (IO_LAST_SEEN,),
    )

    for row_pg in cur_pg.fetchall():
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        cictt_code = row_pg["cictt_code"]  # type: ignore

        try:
            cur_pg.execute(
                """
            DELETE FROM io_aviation_occurrence_categories
             WHERE cictt_code = %s;
            """,
                (cictt_code,),
            )
            if cur_pg.rowcount > 0:
                count_delete += cur_pg.rowcount
                io_utils.progress_msg(f"Deleted cictt_code={cictt_code}")
        except ForeignKeyViolation:
            io_utils.progress_msg(f"Failed to delete cictt_code={cictt_code}")

    conn_pg.commit()

    if count_select > 0:
        io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    if count_delete > 0:
        io_utils.progress_msg(f"Number rows deleted  : {str(count_delete):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_aviation_occurrence_categories_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load country data from a JSON file.
# ------------------------------------------------------------------
def _load_country_data(
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load country data from a JSON file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    filename_json = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_countries_states_json.replace("/", os.sep),
    )

    if not os.path.isfile(filename_json):
        # ERROR.00.934 The country and state data file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_934.replace("{filename}", filename_json)
        )

    count_insert = 0
    count_select = 0
    count_update = 0

    with open(filename_json, "r", encoding=io_glob.FILE_ENCODING_DEFAULT) as input_file:
        input_data = json.load(input_file)

        for record in input_data:
            count_select += 1

            if record["type"] == "country":
                try:
                    cur_pg.execute(
                        """
                    INSERT INTO io_countries (
                           country,country_name,dec_latitude,dec_longitude,first_processed
                           ) VALUES (
                           %s,%s,%s,%s,%s);
                    """,
                        (
                            record["country"],
                            record["country_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                        ),
                    )
                    count_insert += 1
                except UniqueViolation:
                    # pylint: disable=line-too-long
                    cur_pg.execute(
                        """
                    UPDATE io_countries SET
                           country_name = %s,dec_latitude = %s,dec_longitude = %s,last_processed = %s
                     WHERE
                           country = %s
                       AND NOT (country_name = %s AND dec_latitude = %s AND dec_longitude = %s);
                    """,
                        (
                            record["country_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                            record["country"],
                            record["country_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                        ),
                    )
                    count_update += cur_pg.rowcount

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_countries_states_json, cur_pg
    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load NPIAS data from an MS Excel file.
# ------------------------------------------------------------------
def _load_npias_data(us_states) -> list[str]:
    # ------------------------------------------------------------------
    # Start processing NPIAS data.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_faa_npias_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.945 The NPIAS file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_945.replace("{filename}", filename_excel)
        )

    # INFO.00.089 Database table io_airports: Load data from file '{filename}'
    io_utils.progress_msg("-" * 80)
    io_utils.progress_msg(io_glob.INFO_00_089.replace("{filename}", filename_excel))

    count_select = 0
    count_usable = 0

    state_idx = 0
    locid_idx = 3

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    locids: list[str] = []

    # pylint: disable=R0801
    for row in workbook.active:
        count_select += 1
        if not row[state_idx].value in us_states:
            continue

        locids.append(row[locid_idx].value)

        count_usable += 1

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    io_utils.progress_msg(f"Number rows usable   : {str(count_usable):>8}")

    return locids


# ------------------------------------------------------------------
# Load runway data from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def _load_runway_data() -> None:
    io_glob.logger.debug(io_glob.LOGGER_START)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Load airport identifications.
    # ------------------------------------------------------------------
    # INFO.00.088 Database table io_airports: Load the global identifications
    io_utils.progress_msg("-" * 80)
    io_utils.progress_msg(io_glob.INFO_00_088)

    count_select = 0

    cur_pg.execute(
        """
    SELECT global_id
      FROM io_airports
     ORDER BY 1;
    """,
    )

    runway_data: dict[str, tuple[str | None, float | None]] = {}

    for row in cur_pg:
        count_select += 1
        runway_data[row[0]] = (None, None)

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Start processing airport data.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_faa_runways_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.944 The runway file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_944.replace("{filename}", filename_excel)
        )

    # INFO.00.089 Database table io_airports: Load data from file '{filename}'
    io_utils.progress_msg(io_glob.INFO_00_089.replace("{filename}", filename_excel))

    count_select = 0

    airport_id_idx = 2
    comp_code_idx = 7
    dim_uom_idx = 6
    length_idx = 4

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        airport_id = row[airport_id_idx].value
        if not airport_id in runway_data:
            continue

        count_select += 1

        (
            comp_code,
            length,
        ) = runway_data[airport_id]

        new_length = row[length_idx].value

        dim_vom = row[dim_uom_idx].value
        if dim_vom == "M":
            new_length = new_length * 3.28084

        if length is None or new_length > length:
            runway_data[airport_id] = (
                row[comp_code_idx].value,
                new_length,
            )

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load the runway data.
    # ------------------------------------------------------------------

    # INFO.00.090 Database table io_airports: Update the runway data
    io_utils.progress_msg(io_glob.INFO_00_090)

    count_select = 0
    count_update = 0

    # pylint: disable=R0801
    for airport_id, (comp_code, length) in runway_data.items():
        count_select += 1

        if comp_code is None:
            continue

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        cur_pg.execute(
            """
        UPDATE io_airports ia
           SET max_runway_comp_code = %s,
               max_runway_length = %s,
               last_processed = %s
         WHERE ia.global_id = %s
        """,
            (
                comp_code,
                length,
                datetime.now(),
                airport_id,
            ),
        )
        count_update += cur_pg.rowcount

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_sequence_of_events_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load sequence of events sequence data from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-statements
def _load_sequence_of_events() -> None:
    """Load sequence of events sequence data from an MS Excel file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_sequence_of_events_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.938 The sequence of events file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_938.replace("{filename}", filename_excel)
        )

    # INFO.00.076 Database table io_sequence_of_events: Load data from file '{filename}'
    io_utils.progress_msg(
        io_glob.INFO_00_076.replace(
            "{filename}", io_config.settings.download_file_sequence_of_events_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    count_delete = 0
    count_upsert = 0
    count_select = 0

    eventsoe_no_idx = 0
    meaning_idx = 1
    cictt_code_idx = 2

    # ------------------------------------------------------------------
    # Load the sequence of events data.
    # ------------------------------------------------------------------

    cictt_codes = _sql_query_cictt_codes(conn_pg)

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        soe_no = str(row[eventsoe_no_idx].value).rjust(3, "0")
        if soe_no == "eventsoe_no":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        cictt_code = (
            row[cictt_code_idx].value.upper().rstrip()
            if row[cictt_code_idx].value
            else None
        )
        if cictt_code is None or not cictt_code in cictt_codes:
            continue

        meaning = row[meaning_idx].value.rstrip()

        cur_pg.execute(
            """
        INSERT INTO io_sequence_of_events AS isoe (
               soe_no,
               meaning,
               cictt_code,
               first_processed,
               last_seen
               ) VALUES (
               %s,%s,%s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_sequence_of_events_pkey
        DO UPDATE
           SET meaning = %s,
               cictt_code = %s,
               last_processed = %s,
               last_seen = %s
         WHERE isoe.soe_no = %s
        """,
            (
                soe_no,
                meaning,
                cictt_code if cictt_code else None,
                datetime.now(),
                IO_LAST_SEEN,
                meaning,
                cictt_code,
                datetime.now(),
                IO_LAST_SEEN,
                soe_no,
            ),
        )
        count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    # ------------------------------------------------------------------
    # Delete the obsolete data.
    # ------------------------------------------------------------------

    count_select = 0

    # pylint: disable=line-too-long
    cur_pg.execute(
        """
    SELECT soe_no
      FROM io_sequence_of_events
     WHERE last_seen <> %s
        """,
        (IO_LAST_SEEN,),
    )

    for row_pg in cur_pg.fetchall():
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        soe_no = row_pg["soe_no"]  # type: ignore

        try:
            cur_pg.execute(
                """
            DELETE FROM io_sequence_of_events
             WHERE soe_no = %s;
            """,
                (soe_no,),
            )
            if cur_pg.rowcount > 0:
                count_delete += cur_pg.rowcount
                io_utils.progress_msg(f"Deleted soe_no={soe_no}")
        except ForeignKeyViolation:
            io_utils.progress_msg(f"Failed to delete soe_no={soe_no}")

    conn_pg.commit()

    if count_select > 0:
        io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    if count_delete > 0:
        io_utils.progress_msg(f"Number rows deleted  : {str(count_delete):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_sequence_of_events_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load city data from a US city file.
# ------------------------------------------------------------------
# pylint: disable=too-many-locals
def _load_simplemaps_data_cities_from_us_cities(
    conn_pg: connection, cur_pg: cursor
) -> None:
    """Load city data from a US city file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    filename = io_config.settings.download_file_simplemaps_us_cities_xlsx

    if not os.path.isfile(filename):
        # ERROR.00.914 The US city file '{filename}' is missing
        io_utils.terminate_fatal(io_glob.ERROR_00_914.replace("{filename}", filename))

    # INFO.00.027 Database table io_lat_lng: Load city data from file '{filename}'
    io_utils.progress_msg(
        io_glob.INFO_00_027.replace(
            "{filename}", io_config.settings.download_file_simplemaps_us_cities_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    count_upsert = 0
    count_select = 0

    city_idx = 0
    lat_idx = 6
    lng_idx = 7
    state_idx = 2

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        city = row[city_idx].value.upper().rstrip()
        if city == "CITY":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        lat = row[lat_idx].value
        lng = row[lng_idx].value
        state = row[state_idx].value.upper().rstrip()

        # pylint: disable=line-too-long
        cur_pg.execute(
            """
        INSERT INTO io_lat_lng AS ill (
               type,
               country,
               state,
               city,
               dec_latitude,
               dec_longitude,
               source,
               first_processed
               ) VALUES (
               %s,%s,%s,%s,%s,
               %s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
        DO UPDATE
           SET dec_latitude = %s,
               dec_longitude = %s,
               source = %s,
               last_processed = %s
         WHERE ill.type = %s
           AND ill.country = %s
           AND ill.state = %s
           AND ill.city = %s
           AND NOT (
               ill.dec_latitude = %s
           AND ill.dec_longitude = %s
           );
        """,
            (
                io_glob.IO_LAT_LNG_TYPE_CITY,
                io_glob.COUNTRY_USA,
                state,
                city,
                lat,
                lng,
                io_glob.SOURCE_SM_US_CITIES,
                datetime.now(),
                lat,
                lng,
                io_glob.SOURCE_SM_US_CITIES,
                datetime.now(),
                io_glob.IO_LAT_LNG_TYPE_CITY,
                io_glob.COUNTRY_USA,
                state,
                city,
                lat,
                lng,
            ),
        )
        count_upsert += cur_pg.rowcount

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_simplemaps_us_cities_xlsx, cur_pg
    )

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load zip data from a US city file.
# ------------------------------------------------------------------
# pylint: disable=R0801
# pylint: disable=too-many-locals
def _load_simplemaps_data_zips_from_us_cities(
    conn_pg: connection, cur_pg: cursor
) -> None:
    """Load zip data from a US city file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    filename = io_config.settings.download_file_simplemaps_us_cities_xlsx

    if not os.path.isfile(filename):
        # ERROR.00.914 The US city file '{filename}' is missing
        io_utils.terminate_fatal(io_glob.ERROR_00_914.replace("{filename}", filename))

    # INFO.00.039 Database table io_lat_lng: Load zipcode data from file '{filename}'
    io_utils.progress_msg(
        io_glob.INFO_00_039.replace(
            "{filename}", io_config.settings.download_file_simplemaps_us_cities_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    count_insert = 0
    count_select = 0

    city_idx = 0
    lat_idx = 6
    lng_idx = 7
    state_idx = 2
    zipcodes_idx = 15

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    for row in workbook.active:
        city = row[city_idx].value.upper().rstrip()
        if city == "CITY":
            continue

        lat = row[lat_idx].value
        lng = row[lng_idx].value
        state = row[state_idx].value.upper().rstrip()
        zipcodes = list(f"{row[zipcodes_idx].value}".split(" "))

        for zipcode in zipcodes:
            count_select += 1
            if count_select % io_config.settings.database_commit_size == 0:
                conn_pg.commit()
                io_utils.progress_msg(
                    f"Number of rows so far read : {str(count_select):>8}"
                )

            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO io_lat_lng (
                   type,
                   country,
                   state,
                   city,
                   zipcode,
                   dec_latitude,
                   dec_longitude,
                   source,
                   first_processed
                   ) VALUES (
                   %s,%s,%s,%s,%s,
                   %s,%s,%s,%s)
            ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
            DO NOTHING;
            """,
                (
                    io_glob.IO_LAT_LNG_TYPE_ZIPCODE,
                    io_glob.COUNTRY_USA,
                    state,
                    city,
                    zipcode.rstrip(),
                    lat,
                    lng,
                    io_glob.SOURCE_SM_US_CITIES,
                    datetime.now(),
                ),
            )
            count_insert += cur_pg.rowcount

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_simplemaps_us_cities_xlsx, cur_pg
    )

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load zip data from a US zip code file.
# ------------------------------------------------------------------
# pylint: disable=too-many-locals
def _load_simplemaps_data_zips_from_us_zips(
    conn_pg: connection, cur_pg: cursor
) -> None:
    """Load zip data from a US zip code file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    filename = io_config.settings.download_file_simplemaps_us_zips_xlsx

    if not os.path.isfile(filename):
        # ERROR.00.913 The US zip code file '{filename}' is missing
        io_utils.terminate_fatal(io_glob.ERROR_00_913.replace("{filename}", filename))

    # INFO.00.025 Database table io_lat_lng: Load zipcode data rom file '{filename}'
    io_utils.progress_msg(
        io_glob.INFO_00_025.replace(
            "{filename}", io_config.settings.download_file_simplemaps_us_zips_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    count_duplicates = 0
    count_upsert = 0
    count_select = 0

    zip_idx = 0
    lat_idx = 1
    lng_idx = 2
    city_idx = 3
    state_idx = 4

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    for row in workbook.active:
        zipcode = f"{row[zip_idx].value:05}".rstrip()
        if zipcode == "zip00":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        city = row[city_idx].value.upper().rstrip()
        lat = row[lat_idx].value
        lng = row[lng_idx].value
        state = row[state_idx].value.upper().rstrip()

        # pylint: disable=line-too-long
        cur_pg.execute(
            """
        INSERT INTO io_lat_lng AS ill (
               type,
               country,
               state,
               city,
               zipcode,
               dec_latitude,
               dec_longitude,
               source,
               first_processed
               ) VALUES (
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
        DO UPDATE
           SET dec_latitude = %s,
               dec_longitude = %s,
               source = %s,
               last_processed = %s
         WHERE ill.type = %s
           AND ill.country = %s
           AND ill.state = %s
           AND ill.city = %s
           AND ill.zipcode = %s
           AND NOT (
               ill.dec_latitude = %s
           AND ill.dec_longitude = %s
           );
        """,
            (
                io_glob.IO_LAT_LNG_TYPE_ZIPCODE,
                io_glob.COUNTRY_USA,
                state,
                city,
                zipcode,
                lat,
                lng,
                io_glob.SOURCE_SM_US_ZIP_CODES,
                datetime.now(),
                lat,
                lng,
                io_glob.SOURCE_SM_US_ZIP_CODES,
                datetime.now(),
                io_glob.IO_LAT_LNG_TYPE_ZIPCODE,
                io_glob.COUNTRY_USA,
                state,
                city,
                zipcode,
                lat,
                lng,
            ),
        )
        count_upsert += cur_pg.rowcount

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_simplemaps_us_zips_xlsx, cur_pg
    )

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    if count_duplicates > 0:
        io_utils.progress_msg(f"Number rows duplicate: {str(count_duplicates):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load state data from a JSON file.
# ------------------------------------------------------------------
def _load_state_data(
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load state data from a JSON file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    filename_json = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_countries_states_json.replace("/", os.sep),
    )

    if not os.path.isfile(filename_json):
        # ERROR.00.934 The country and state data file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_934.replace("{filename}", filename_json)
        )

    count_insert = 0
    count_select = 0
    count_update = 0

    with open(filename_json, "r", encoding=io_glob.FILE_ENCODING_DEFAULT) as input_file:
        input_data = json.load(input_file)

        for record in input_data:
            count_select += 1

            if record["type"] == "state":
                try:
                    cur_pg.execute(
                        """
                    INSERT INTO io_states (
                           country,state, state_name,dec_latitude,dec_longitude,first_processed
                           ) VALUES (%s,%s,%s,%s,%s,%s);
                    """,
                        (
                            record["country"],
                            record["state"],
                            record["state_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                        ),
                    )
                    count_insert += 1
                except UniqueViolation:
                    cur_pg.execute(
                        """
                    UPDATE io_states SET
                           state_name = %s,dec_latitude = %s,dec_longitude = %s,last_processed = %s
                     WHERE
                           country = %s AND state = %s
                       AND NOT (state_name = %s AND dec_latitude = %s AND dec_longitude = %s);
                    """,
                        (
                            record["state_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                            record["country"],
                            record["state"],
                            record["state_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                        ),
                    )
                    count_update += cur_pg.rowcount

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_countries_states_json, cur_pg
    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table aircraft.
# ------------------------------------------------------------------
def _load_table_aircraft(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table aircraft."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    # pylint: disable=R0801
    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO aircraft (
                   ev_id,aircraft_key,regis_no,ntsb_no,acft_missing,far_part,flt_plan_filed,flight_plan_activated,damage,acft_fire,acft_expl,acft_make,
                   acft_model,acft_series,acft_serial_no,cert_max_gr_wt,acft_category,acft_reg_cls,homebuilt,fc_seats,cc_seats,pax_seats,total_seats,
                   num_eng,fixed_retractable,type_last_insp,date_last_insp,afm_hrs_last_insp,afm_hrs,elt_install,elt_oper,elt_aided_loc_ev,elt_type,
                   owner_acft,owner_street,owner_city,owner_state,owner_country,owner_zip,oper_individual_name,oper_name,oper_same,oper_dba,oper_addr_same,
                   oper_street,oper_city,oper_state,oper_country,oper_zip,oper_code,certs_held,oprtng_cert,oper_cert,oper_cert_num,oper_sched,oper_dom_int,
                   oper_pax_cargo,type_fly,second_pilot,dprt_pt_same_ev,dprt_apt_id,dprt_city,dprt_state,dprt_country,dprt_time,dprt_timezn,dest_same_local,
                   dest_apt_id,dest_city,dest_state,dest_country,phase_flt_spec,report_to_icao,evacuation,lchg_date,lchg_userid,afm_hrs_since,rwy_num,
                   rwy_len,rwy_width,site_seeing,air_medical,med_type_flight,acft_year,fuel_on_board,commercial_space_flight,unmanned,ifr_equipped_cert,
                   elt_mounted_aircraft,elt_connected_antenna,elt_manufacturer,elt_model,elt_reason_other,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                   %s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.regis_no,
                    row_mdb.ntsb_no,
                    row_mdb.acft_missing,
                    row_mdb.far_part,
                    row_mdb.flt_plan_filed,
                    row_mdb.flight_plan_activated,
                    row_mdb.damage,
                    row_mdb.acft_fire,
                    row_mdb.acft_expl,
                    row_mdb.acft_make,
                    row_mdb.acft_model,
                    row_mdb.acft_series,
                    row_mdb.acft_serial_no,
                    row_mdb.cert_max_gr_wt,
                    row_mdb.acft_category,
                    row_mdb.acft_reg_cls,
                    row_mdb.homebuilt,
                    row_mdb.fc_seats,
                    row_mdb.cc_seats,
                    row_mdb.pax_seats,
                    row_mdb.total_seats,
                    row_mdb.num_eng,
                    row_mdb.fixed_retractable,
                    row_mdb.type_last_insp,
                    row_mdb.date_last_insp,
                    row_mdb.afm_hrs_last_insp,
                    row_mdb.afm_hrs,
                    row_mdb.elt_install,
                    row_mdb.elt_oper,
                    row_mdb.elt_aided_loc_ev,
                    row_mdb.elt_type,
                    row_mdb.owner_acft,
                    row_mdb.owner_street,
                    row_mdb.owner_city,
                    row_mdb.owner_state,
                    row_mdb.owner_country,
                    row_mdb.owner_zip,
                    row_mdb.oper_individual_name,
                    row_mdb.oper_name,
                    row_mdb.oper_same,
                    row_mdb.oper_dba,
                    row_mdb.oper_addr_same,
                    row_mdb.oper_street,
                    row_mdb.oper_city,
                    row_mdb.oper_state,
                    row_mdb.oper_country,
                    row_mdb.oper_zip,
                    row_mdb.oper_code,
                    row_mdb.certs_held,
                    row_mdb.oprtng_cert,
                    row_mdb.oper_cert,
                    row_mdb.oper_cert_num,
                    row_mdb.oper_sched,
                    row_mdb.oper_dom_int,
                    row_mdb.oper_pax_cargo,
                    row_mdb.type_fly,
                    row_mdb.second_pilot,
                    row_mdb.dprt_pt_same_ev,
                    row_mdb.dprt_apt_id,
                    row_mdb.dprt_city,
                    row_mdb.dprt_state,
                    row_mdb.dprt_country,
                    row_mdb.dprt_time,
                    row_mdb.dprt_timezn,
                    row_mdb.dest_same_local,
                    row_mdb.dest_apt_id,
                    row_mdb.dest_city,
                    row_mdb.dest_state,
                    row_mdb.dest_country,
                    row_mdb.phase_flt_spec,
                    row_mdb.report_to_icao,
                    row_mdb.evacuation,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    row_mdb.afm_hrs_since,
                    row_mdb.rwy_num,
                    row_mdb.rwy_len,
                    row_mdb.rwy_width,
                    row_mdb.site_seeing,
                    row_mdb.air_medical,
                    row_mdb.med_type_flight,
                    row_mdb.acft_year,
                    row_mdb.fuel_on_board,
                    row_mdb.commercial_space_flight,
                    row_mdb.unmanned,
                    row_mdb.ifr_equipped_cert,
                    row_mdb.elt_mounted_aircraft,
                    row_mdb.elt_connected_antenna,
                    row_mdb.elt_manufacturer,
                    row_mdb.elt_model,
                    row_mdb.elt_reason_other,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE aircraft
               SET regis_no = %s,ntsb_no = %s,acft_missing = %s,far_part = %s,flt_plan_filed = %s,flight_plan_activated = %s,damage = %s,acft_fire = %s,
                   acft_expl = %s,acft_make = %s,acft_model = %s,acft_series = %s,acft_serial_no = %s,cert_max_gr_wt = %s,acft_category = %s,
                   acft_reg_cls = %s,homebuilt = %s,fc_seats = %s,cc_seats = %s,pax_seats = %s,total_seats = %s,num_eng = %s,fixed_retractable = %s,
                   type_last_insp = %s,date_last_insp = %s,afm_hrs_last_insp = %s,afm_hrs = %s,elt_install = %s,elt_oper = %s,elt_aided_loc_ev = %s,
                   elt_type = %s,owner_acft = %s,owner_street = %s,owner_city = %s,owner_state = %s,owner_country = %s,owner_zip = %s,
                   oper_individual_name = %s,oper_name = %s,oper_same = %s,oper_dba = %s,oper_addr_same = %s,oper_street = %s,oper_city = %s,
                   oper_state = %s,oper_country = %s,oper_zip = %s,oper_code = %s,certs_held = %s,oprtng_cert = %s,oper_cert = %s,oper_cert_num = %s,
                   oper_sched = %s,oper_dom_int = %s,oper_pax_cargo = %s,type_fly = %s,second_pilot = %s,dprt_pt_same_ev = %s,dprt_apt_id = %s,
                   dprt_city = %s,dprt_state = %s,dprt_country = %s,dprt_time = %s,dprt_timezn = %s,dest_same_local = %s,dest_apt_id = %s,dest_city = %s,
                   dest_state = %s,dest_country = %s,phase_flt_spec = %s,report_to_icao = %s,evacuation = %s,lchg_date = %s,lchg_userid = %s,
                   afm_hrs_since = %s,rwy_num = %s,rwy_len = %s,rwy_width = %s,site_seeing = %s,air_medical = %s,med_type_flight = %s,acft_year = %s,
                   fuel_on_board = %s,commercial_space_flight = %s,unmanned = %s,ifr_equipped_cert = %s,elt_mounted_aircraft = %s,elt_connected_antenna = %s,
                   elt_manufacturer = %s,elt_model = %s,elt_reason_other = %s,io_last_seen_ntsb = %s
             WHERE aircraft_key = %s AND ev_id = %s
               AND NOT (regis_no = %s AND ntsb_no = %s AND acft_missing = %s AND far_part = %s AND flt_plan_filed = %s AND flight_plan_activated = %s AND
                        damage = %s AND acft_fire = %s AND acft_expl = %s AND acft_make = %s AND acft_model = %s AND acft_series = %s AND
                        acft_serial_no = %s AND cert_max_gr_wt = %s AND acft_category = %s AND acft_reg_cls = %s AND homebuilt = %s AND fc_seats = %s AND
                        cc_seats = %s AND pax_seats = %s AND total_seats = %s AND num_eng = %s AND fixed_retractable = %s AND type_last_insp = %s AND
                        date_last_insp = %s AND afm_hrs_last_insp = %s AND afm_hrs = %s AND elt_install = %s AND elt_oper = %s AND elt_aided_loc_ev = %s AND
                        elt_type = %s AND owner_acft = %s AND owner_street = %s AND owner_city = %s AND owner_state = %s AND owner_country = %s AND
                        owner_zip = %s AND oper_individual_name = %s AND oper_name = %s AND oper_same = %s AND oper_dba = %s AND oper_addr_same = %s AND
                        oper_street = %s AND oper_city = %s AND oper_state = %s AND oper_country = %s AND oper_zip = %s AND oper_code = %s AND
                        certs_held = %s AND oprtng_cert = %s AND oper_cert = %s AND oper_cert_num = %s AND oper_sched = %s AND oper_dom_int = %s AND
                        oper_pax_cargo = %s AND type_fly = %s AND second_pilot = %s AND dprt_pt_same_ev = %s AND dprt_apt_id = %s AND dprt_city = %s AND
                        dprt_state = %s AND dprt_country = %s AND dprt_time = %s AND dprt_timezn = %s AND dest_same_local = %s AND dest_apt_id = %s AND
                        dest_city = %s AND dest_state = %s AND dest_country = %s AND phase_flt_spec = %s AND report_to_icao = %s AND evacuation = %s AND
                        lchg_date = %s AND lchg_userid = %s AND afm_hrs_since = %s AND rwy_num = %s AND rwy_len = %s AND rwy_width = %s AND site_seeing = %s AND
                        air_medical = %s AND med_type_flight = %s AND acft_year = %s AND fuel_on_board = %s AND commercial_space_flight = %s AND
                        unmanned = %s AND ifr_equipped_cert = %s AND elt_mounted_aircraft = %s AND elt_connected_antenna = %s AND elt_manufacturer = %s AND
                        elt_model = %s AND elt_reason_other = %s);
            """,
                (
                    row_mdb.regis_no,
                    row_mdb.ntsb_no,
                    row_mdb.acft_missing,
                    row_mdb.far_part,
                    row_mdb.flt_plan_filed,
                    row_mdb.flight_plan_activated,
                    row_mdb.damage,
                    row_mdb.acft_fire,
                    row_mdb.acft_expl,
                    row_mdb.acft_make,
                    row_mdb.acft_model,
                    row_mdb.acft_series,
                    row_mdb.acft_serial_no,
                    row_mdb.cert_max_gr_wt,
                    row_mdb.acft_category,
                    row_mdb.acft_reg_cls,
                    row_mdb.homebuilt,
                    row_mdb.fc_seats,
                    row_mdb.cc_seats,
                    row_mdb.pax_seats,
                    row_mdb.total_seats,
                    row_mdb.num_eng,
                    row_mdb.fixed_retractable,
                    row_mdb.type_last_insp,
                    row_mdb.date_last_insp,
                    row_mdb.afm_hrs_last_insp,
                    row_mdb.afm_hrs,
                    row_mdb.elt_install,
                    row_mdb.elt_oper,
                    row_mdb.elt_aided_loc_ev,
                    row_mdb.elt_type,
                    row_mdb.owner_acft,
                    row_mdb.owner_street,
                    row_mdb.owner_city,
                    row_mdb.owner_state,
                    row_mdb.owner_country,
                    row_mdb.owner_zip,
                    row_mdb.oper_individual_name,
                    row_mdb.oper_name,
                    row_mdb.oper_same,
                    row_mdb.oper_dba,
                    row_mdb.oper_addr_same,
                    row_mdb.oper_street,
                    row_mdb.oper_city,
                    row_mdb.oper_state,
                    row_mdb.oper_country,
                    row_mdb.oper_zip,
                    row_mdb.oper_code,
                    row_mdb.certs_held,
                    row_mdb.oprtng_cert,
                    row_mdb.oper_cert,
                    row_mdb.oper_cert_num,
                    row_mdb.oper_sched,
                    row_mdb.oper_dom_int,
                    row_mdb.oper_pax_cargo,
                    row_mdb.type_fly,
                    row_mdb.second_pilot,
                    row_mdb.dprt_pt_same_ev,
                    row_mdb.dprt_apt_id,
                    row_mdb.dprt_city,
                    row_mdb.dprt_state,
                    row_mdb.dprt_country,
                    row_mdb.dprt_time,
                    row_mdb.dprt_timezn,
                    row_mdb.dest_same_local,
                    row_mdb.dest_apt_id,
                    row_mdb.dest_city,
                    row_mdb.dest_state,
                    row_mdb.dest_country,
                    row_mdb.phase_flt_spec,
                    row_mdb.report_to_icao,
                    row_mdb.evacuation,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    row_mdb.afm_hrs_since,
                    row_mdb.rwy_num,
                    row_mdb.rwy_len,
                    row_mdb.rwy_width,
                    row_mdb.site_seeing,
                    row_mdb.air_medical,
                    row_mdb.med_type_flight,
                    row_mdb.acft_year,
                    row_mdb.fuel_on_board,
                    row_mdb.commercial_space_flight,
                    row_mdb.unmanned,
                    row_mdb.ifr_equipped_cert,
                    row_mdb.elt_mounted_aircraft,
                    row_mdb.elt_connected_antenna,
                    row_mdb.elt_manufacturer,
                    row_mdb.elt_model,
                    row_mdb.elt_reason_other,
                    IO_LAST_SEEN,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.regis_no,
                    row_mdb.ntsb_no,
                    row_mdb.acft_missing,
                    row_mdb.far_part,
                    row_mdb.flt_plan_filed,
                    row_mdb.flight_plan_activated,
                    row_mdb.damage,
                    row_mdb.acft_fire,
                    row_mdb.acft_expl,
                    row_mdb.acft_make,
                    row_mdb.acft_model,
                    row_mdb.acft_series,
                    row_mdb.acft_serial_no,
                    row_mdb.cert_max_gr_wt,
                    row_mdb.acft_category,
                    row_mdb.acft_reg_cls,
                    row_mdb.homebuilt,
                    row_mdb.fc_seats,
                    row_mdb.cc_seats,
                    row_mdb.pax_seats,
                    row_mdb.total_seats,
                    row_mdb.num_eng,
                    row_mdb.fixed_retractable,
                    row_mdb.type_last_insp,
                    row_mdb.date_last_insp,
                    row_mdb.afm_hrs_last_insp,
                    row_mdb.afm_hrs,
                    row_mdb.elt_install,
                    row_mdb.elt_oper,
                    row_mdb.elt_aided_loc_ev,
                    row_mdb.elt_type,
                    row_mdb.owner_acft,
                    row_mdb.owner_street,
                    row_mdb.owner_city,
                    row_mdb.owner_state,
                    row_mdb.owner_country,
                    row_mdb.owner_zip,
                    row_mdb.oper_individual_name,
                    row_mdb.oper_name,
                    row_mdb.oper_same,
                    row_mdb.oper_dba,
                    row_mdb.oper_addr_same,
                    row_mdb.oper_street,
                    row_mdb.oper_city,
                    row_mdb.oper_state,
                    row_mdb.oper_country,
                    row_mdb.oper_zip,
                    row_mdb.oper_code,
                    row_mdb.certs_held,
                    row_mdb.oprtng_cert,
                    row_mdb.oper_cert,
                    row_mdb.oper_cert_num,
                    row_mdb.oper_sched,
                    row_mdb.oper_dom_int,
                    row_mdb.oper_pax_cargo,
                    row_mdb.type_fly,
                    row_mdb.second_pilot,
                    row_mdb.dprt_pt_same_ev,
                    row_mdb.dprt_apt_id,
                    row_mdb.dprt_city,
                    row_mdb.dprt_state,
                    row_mdb.dprt_country,
                    row_mdb.dprt_time,
                    row_mdb.dprt_timezn,
                    row_mdb.dest_same_local,
                    row_mdb.dest_apt_id,
                    row_mdb.dest_city,
                    row_mdb.dest_state,
                    row_mdb.dest_country,
                    row_mdb.phase_flt_spec,
                    row_mdb.report_to_icao,
                    row_mdb.evacuation,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    row_mdb.afm_hrs_since,
                    row_mdb.rwy_num,
                    row_mdb.rwy_len,
                    row_mdb.rwy_width,
                    row_mdb.site_seeing,
                    row_mdb.air_medical,
                    row_mdb.med_type_flight,
                    row_mdb.acft_year,
                    row_mdb.fuel_on_board,
                    row_mdb.commercial_space_flight,
                    row_mdb.unmanned,
                    row_mdb.ifr_equipped_cert,
                    row_mdb.elt_mounted_aircraft,
                    row_mdb.elt_connected_antenna,
                    row_mdb.elt_manufacturer,
                    row_mdb.elt_model,
                    row_mdb.elt_reason_other,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table dt_aircraft.
# ------------------------------------------------------------------
def _load_table_dt_aircraft(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table dt_aircraft."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            cur_pg.execute(
                """
            INSERT INTO dt_aircraft (
                   ev_id,
                   aircraft_key,
                   col_name,
                   code,
                   lchg_date,
                   lchg_userid,
                   io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,
                   %s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.col_name,
                    row_mdb.code,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            UPDATE dt_aircraft
               SET lchg_date = %s,
                   lchg_userid = %s,
                   io_last_seen_ntsb = %s
             WHERE code = %s 
               AND col_name = %s 
               AND aircraft_key = %s 
               AND ev_id = %s
               AND NOT (lchg_date = %s 
                    AND lchg_userid = %s);
            """,
                (
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.code,
                    row_mdb.col_name,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"col_name={row_mdb.col_name} "
                        + f"code={row_mdb.code}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table dt_events.
# ------------------------------------------------------------------
def _load_table_dt_events(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table dt_events."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            cur_pg.execute(
                """
            INSERT INTO dt_events (
                   ev_id,
                   col_name,
                   code,
                   lchg_date,
                   lchg_userid,
                   io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,
                   %s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.col_name,
                    row_mdb.code,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE dt_events
               SET lchg_date = %s,
                   lchg_userid = %s,
                   io_last_seen_ntsb = %s
             WHERE code = %s 
               AND col_name = %s 
               AND ev_id = %s
               AND NOT (lchg_date = %s 
                    AND lchg_userid = %s);
            """,
                (
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.code,
                    row_mdb.col_name,
                    row_mdb.ev_id,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"col_name={row_mdb.col_name} "
                        + f"code={row_mdb.code}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table dt_flight_crew.
# ------------------------------------------------------------------
def _load_table_dt_flight_crew(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table dt_flight_crew."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            cur_pg.execute(
                """
            INSERT INTO dt_flight_crew (
                   ev_id,aircraft_key,crew_no,col_name,code,lchg_date,lchg_userid,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.crew_no,
                    row_mdb.col_name,
                    row_mdb.code,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            UPDATE dt_flight_crew
               SET lchg_date = %s,lchg_userid = %s,io_last_seen_ntsb = %s
             WHERE code = %s AND col_name = %s AND crew_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.code,
                    row_mdb.col_name,
                    row_mdb.crew_no,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"Aircraft_Key={row_mdb.Aircraft_Key} "
                        + f"crew_no={row_mdb.crew_no} "
                        + f"col_name={row_mdb.col_name} "
                        + f"code={row_mdb.code}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table engines.
# ------------------------------------------------------------------
def _load_table_engines(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table engines."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO engines (
                   ev_id,aircraft_key,eng_no,eng_type,eng_mfgr,eng_model,power_units,hp_or_lbs,lchg_userid,lchg_date,carb_fuel_injection,propeller_type,
                   propeller_make,propeller_model,eng_time_total,eng_time_last_insp,eng_time_overhaul,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.eng_no,
                    row_mdb.eng_type,
                    row_mdb.eng_mfgr,
                    row_mdb.eng_model,
                    row_mdb.power_units,
                    row_mdb.hp_or_lbs,
                    row_mdb.lchg_userid,
                    row_mdb.lchg_date,
                    row_mdb.carb_fuel_injection,
                    row_mdb.propeller_type,
                    row_mdb.propeller_make,
                    row_mdb.propeller_model,
                    row_mdb.eng_time_total,
                    row_mdb.eng_time_last_insp,
                    row_mdb.eng_time_overhaul,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE engines
               SET eng_type = %s,eng_mfgr = %s,eng_model = %s,power_units = %s,hp_or_lbs = %s,lchg_userid = %s,lchg_date = %s,carb_fuel_injection = %s,
                   propeller_type = %s,propeller_make = %s,propeller_model = %s,eng_time_total = %s,eng_time_last_insp = %s,eng_time_overhaul = %s,
                   io_last_seen_ntsb = %s
             WHERE eng_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (eng_type = %s AND eng_mfgr = %s AND eng_model = %s AND power_units = %s AND hp_or_lbs = %s AND lchg_userid = %s AND
                        lchg_date = %s AND carb_fuel_injection = %s AND propeller_type = %s AND propeller_make = %s AND propeller_model = %s AND
                        eng_time_total = %s AND eng_time_last_insp = %s AND eng_time_overhaul = %s);
            """,
                (
                    row_mdb.eng_type,
                    row_mdb.eng_mfgr,
                    row_mdb.eng_model,
                    row_mdb.power_units,
                    row_mdb.hp_or_lbs,
                    row_mdb.lchg_userid,
                    row_mdb.lchg_date,
                    row_mdb.carb_fuel_injection,
                    row_mdb.propeller_type,
                    row_mdb.propeller_make,
                    row_mdb.propeller_model,
                    row_mdb.eng_time_total,
                    row_mdb.eng_time_last_insp,
                    row_mdb.eng_time_overhaul,
                    IO_LAST_SEEN,
                    row_mdb.eng_no,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.eng_type,
                    row_mdb.eng_mfgr,
                    row_mdb.eng_model,
                    row_mdb.power_units,
                    row_mdb.hp_or_lbs,
                    row_mdb.lchg_userid,
                    row_mdb.lchg_date,
                    row_mdb.carb_fuel_injection,
                    row_mdb.propeller_type,
                    row_mdb.propeller_make,
                    row_mdb.propeller_model,
                    row_mdb.eng_time_total,
                    row_mdb.eng_time_last_insp,
                    row_mdb.eng_time_overhaul,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"eng_no={row_mdb.eng_no}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table events.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
def _load_table_events(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table events."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            if msaccess == io_glob.MSACCESS_PRE2008:
                # pylint: disable=line-too-long
                cur_pg.execute(
                    """
                INSERT INTO events (
                       ev_id,ntsb_no,ev_type,ev_date,ev_dow,ev_time,ev_tmzn,ev_city,ev_state,ev_country,ev_site_zipcode,ev_year,ev_month,mid_air,
                       on_ground_collision,latitude,longitude,latlong_acq,apt_name,ev_nr_apt_id,ev_nr_apt_loc,apt_dist,apt_dir,apt_elev,wx_brief_comp,
                       wx_src_iic,wx_obs_time,wx_obs_dir,wx_obs_fac_id,wx_obs_elev,wx_obs_dist,wx_obs_tmzn,light_cond,sky_cond_nonceil,sky_nonceil_ht,
                       sky_ceil_ht,sky_cond_ceil,vis_rvr,vis_rvv,vis_sm,wx_temp,wx_dew_pt,wind_dir_deg,wind_dir_ind,wind_vel_kts,wind_vel_ind,gust_ind,
                       gust_kts,altimeter,wx_dens_alt,wx_int_precip,metar,ev_highest_injury,inj_f_grnd,inj_m_grnd,inj_s_grnd,inj_tot_f,inj_tot_m,inj_tot_n,
                       inj_tot_s,inj_tot_t,invest_agy,ntsb_docket,ntsb_notf_from,ntsb_notf_date,ntsb_notf_tm,fiche_number,lchg_date,lchg_userid,wx_cond_basic,
                       faa_dist_office,io_last_seen_ntsb
                       ) VALUES (
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
                """,
                    (
                        row_mdb.ev_id,
                        row_mdb.ntsb_no,
                        row_mdb.ev_type,
                        row_mdb.ev_date,
                        row_mdb.ev_dow,
                        row_mdb.ev_time,
                        row_mdb.ev_tmzn,
                        None
                        if row_mdb.ev_city is None
                        else row_mdb.ev_city.rstrip()
                        if len(row_mdb.ev_city.rstrip()) > 0
                        else None,
                        row_mdb.ev_state,
                        row_mdb.ev_country,
                        None
                        if row_mdb.ev_site_zipcode is None
                        else row_mdb.ev_site_zipcode.rstrip()
                        if len(row_mdb.ev_site_zipcode.rstrip()) > 0
                        else None,
                        row_mdb.ev_year,
                        row_mdb.ev_month,
                        row_mdb.mid_air,
                        row_mdb.on_ground_collision,
                        None
                        if row_mdb.latitude is None
                        else row_mdb.latitude.rstrip()
                        if len(row_mdb.latitude.rstrip()) > 0
                        else None,
                        None
                        if row_mdb.longitude is None
                        else row_mdb.longitude.rstrip()
                        if len(row_mdb.longitude.rstrip()) > 0
                        else None,
                        row_mdb.latlong_acq,
                        row_mdb.apt_name,
                        row_mdb.ev_nr_apt_id,
                        row_mdb.ev_nr_apt_loc,
                        row_mdb.apt_dist,
                        row_mdb.apt_dir,
                        row_mdb.apt_elev,
                        row_mdb.wx_brief_comp,
                        row_mdb.wx_src_iic,
                        row_mdb.wx_obs_time,
                        row_mdb.wx_obs_dir,
                        row_mdb.wx_obs_fac_id,
                        row_mdb.wx_obs_elev,
                        row_mdb.wx_obs_dist,
                        row_mdb.wx_obs_tmzn,
                        row_mdb.light_cond,
                        row_mdb.sky_cond_nonceil,
                        row_mdb.sky_nonceil_ht,
                        row_mdb.sky_ceil_ht,
                        row_mdb.sky_cond_ceil,
                        row_mdb.vis_rvr,
                        row_mdb.vis_rvv,
                        row_mdb.vis_sm,
                        row_mdb.wx_temp,
                        row_mdb.wx_dew_pt,
                        row_mdb.wind_dir_deg,
                        row_mdb.wind_dir_ind,
                        row_mdb.wind_vel_kts,
                        row_mdb.wind_vel_ind,
                        row_mdb.gust_ind,
                        row_mdb.gust_kts,
                        row_mdb.altimeter,
                        row_mdb.wx_dens_alt,
                        row_mdb.wx_int_precip,
                        row_mdb.metar,
                        row_mdb.ev_highest_injury,
                        row_mdb.inj_f_grnd,
                        row_mdb.inj_m_grnd,
                        row_mdb.inj_s_grnd,
                        row_mdb.inj_tot_f,
                        row_mdb.inj_tot_m,
                        row_mdb.inj_tot_n,
                        row_mdb.inj_tot_s,
                        row_mdb.inj_tot_t,
                        row_mdb.invest_agy,
                        row_mdb.ntsb_docket,
                        row_mdb.ntsb_notf_from,
                        row_mdb.ntsb_notf_date,
                        row_mdb.ntsb_notf_tm,
                        row_mdb.fiche_number,
                        row_mdb.lchg_date,
                        row_mdb.lchg_userid,
                        row_mdb.wx_cond_basic,
                        row_mdb.faa_dist_office,
                        IO_LAST_SEEN,
                    ),
                )
            else:
                # pylint: disable=line-too-long
                cur_pg.execute(
                    """
                INSERT INTO events (
                       ev_id,ntsb_no,ev_type,ev_date,ev_dow,ev_time,ev_tmzn,ev_city,ev_state,ev_country,ev_site_zipcode,ev_year,ev_month,mid_air,
                       on_ground_collision,latitude,longitude,latlong_acq,apt_name,ev_nr_apt_id,ev_nr_apt_loc,apt_dist,apt_dir,apt_elev,wx_brief_comp,
                       wx_src_iic,wx_obs_time,wx_obs_dir,wx_obs_fac_id,wx_obs_elev,wx_obs_dist,wx_obs_tmzn,light_cond,sky_cond_nonceil,sky_nonceil_ht,
                       sky_ceil_ht,sky_cond_ceil,vis_rvr,vis_rvv,vis_sm,wx_temp,wx_dew_pt,wind_dir_deg,wind_dir_ind,wind_vel_kts,wind_vel_ind,gust_ind,
                       gust_kts,altimeter,wx_dens_alt,wx_int_precip,metar,ev_highest_injury,inj_f_grnd,inj_m_grnd,inj_s_grnd,inj_tot_f,inj_tot_m,inj_tot_n,
                       inj_tot_s,inj_tot_t,invest_agy,ntsb_docket,ntsb_notf_from,ntsb_notf_date,ntsb_notf_tm,fiche_number,lchg_date,lchg_userid,wx_cond_basic,
                       faa_dist_office,dec_latitude,dec_longitude,io_last_seen_ntsb
                       ) VALUES (
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
                """,
                    (
                        row_mdb.ev_id,
                        row_mdb.ntsb_no,
                        row_mdb.ev_type,
                        row_mdb.ev_date,
                        row_mdb.ev_dow,
                        row_mdb.ev_time,
                        row_mdb.ev_tmzn,
                        None
                        if row_mdb.ev_city is None
                        else row_mdb.ev_city.rstrip()
                        if len(row_mdb.ev_city.rstrip()) > 0
                        else None,
                        row_mdb.ev_state,
                        row_mdb.ev_country,
                        None
                        if row_mdb.ev_site_zipcode is None
                        else row_mdb.ev_site_zipcode.rstrip()
                        if len(row_mdb.ev_site_zipcode.rstrip()) > 0
                        else None,
                        row_mdb.ev_year,
                        row_mdb.ev_month,
                        row_mdb.mid_air,
                        row_mdb.on_ground_collision,
                        None
                        if row_mdb.latitude is None
                        else row_mdb.latitude.rstrip()
                        if len(row_mdb.latitude.rstrip()) > 0
                        else None,
                        None
                        if row_mdb.longitude is None
                        else row_mdb.longitude.rstrip()
                        if len(row_mdb.longitude.rstrip()) > 0
                        else None,
                        row_mdb.latlong_acq,
                        row_mdb.apt_name,
                        row_mdb.ev_nr_apt_id,
                        row_mdb.ev_nr_apt_loc,
                        row_mdb.apt_dist,
                        row_mdb.apt_dir,
                        row_mdb.apt_elev,
                        row_mdb.wx_brief_comp,
                        row_mdb.wx_src_iic,
                        row_mdb.wx_obs_time,
                        row_mdb.wx_obs_dir,
                        row_mdb.wx_obs_fac_id,
                        row_mdb.wx_obs_elev,
                        row_mdb.wx_obs_dist,
                        row_mdb.wx_obs_tmzn,
                        row_mdb.light_cond,
                        row_mdb.sky_cond_nonceil,
                        row_mdb.sky_nonceil_ht,
                        row_mdb.sky_ceil_ht,
                        row_mdb.sky_cond_ceil,
                        row_mdb.vis_rvr,
                        row_mdb.vis_rvv,
                        row_mdb.vis_sm,
                        row_mdb.wx_temp,
                        row_mdb.wx_dew_pt,
                        row_mdb.wind_dir_deg,
                        row_mdb.wind_dir_ind,
                        row_mdb.wind_vel_kts,
                        row_mdb.wind_vel_ind,
                        row_mdb.gust_ind,
                        row_mdb.gust_kts,
                        row_mdb.altimeter,
                        row_mdb.wx_dens_alt,
                        row_mdb.wx_int_precip,
                        row_mdb.metar,
                        row_mdb.ev_highest_injury,
                        row_mdb.inj_f_grnd,
                        row_mdb.inj_m_grnd,
                        row_mdb.inj_s_grnd,
                        row_mdb.inj_tot_f,
                        row_mdb.inj_tot_m,
                        row_mdb.inj_tot_n,
                        row_mdb.inj_tot_s,
                        row_mdb.inj_tot_t,
                        row_mdb.invest_agy,
                        row_mdb.ntsb_docket,
                        row_mdb.ntsb_notf_from,
                        row_mdb.ntsb_notf_date,
                        row_mdb.ntsb_notf_tm,
                        row_mdb.fiche_number,
                        row_mdb.lchg_date,
                        row_mdb.lchg_userid,
                        row_mdb.wx_cond_basic,
                        row_mdb.faa_dist_office,
                        row_mdb.dec_latitude,
                        row_mdb.dec_longitude,
                        IO_LAST_SEEN,
                    ),
                )
            count_insert += 1
            if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                io_utils.progress_msg(
                    f"Inserted ev_id={row_mdb.ev_id} ev_year={row_mdb.ev_year}"
                )
        except UniqueViolation:
            if msaccess == io_glob.MSACCESS_PRE2008:
                # pylint: disable=line-too-long
                cur_pg.execute(
                    """
                UPDATE events
                   SET ntsb_no = %s,ev_type = %s,ev_date = %s,ev_dow = %s,ev_time = %s,ev_tmzn = %s,ev_city = %s,ev_state = %s,ev_country = %s,
                       ev_site_zipcode = %s,ev_year = %s,ev_month = %s,mid_air = %s,on_ground_collision = %s,latitude = %s,longitude = %s,latlong_acq = %s,
                       apt_name = %s,ev_nr_apt_id = %s,ev_nr_apt_loc = %s,apt_dist = %s,apt_dir = %s,apt_elev = %s,wx_brief_comp = %s,wx_src_iic = %s,
                       wx_obs_time = %s,wx_obs_dir = %s,wx_obs_fac_id = %s,wx_obs_elev = %s,wx_obs_dist = %s,wx_obs_tmzn = %s,light_cond = %s,
                       sky_cond_nonceil = %s,sky_nonceil_ht = %s,sky_ceil_ht = %s,sky_cond_ceil = %s,vis_rvr = %s,vis_rvv = %s,vis_sm = %s,wx_temp = %s,
                       wx_dew_pt = %s,wind_dir_deg = %s,wind_dir_ind = %s,wind_vel_kts = %s,wind_vel_ind = %s,gust_ind = %s,gust_kts = %s,altimeter = %s,
                       wx_dens_alt = %s,wx_int_precip = %s,metar = %s,ev_highest_injury = %s,inj_f_grnd = %s,inj_m_grnd = %s,inj_s_grnd = %s,inj_tot_f = %s,
                       inj_tot_m = %s,inj_tot_n = %s,inj_tot_s = %s,inj_tot_t = %s,invest_agy = %s,ntsb_docket = %s,ntsb_notf_from = %s,ntsb_notf_date = %s,
                       ntsb_notf_tm = %s,fiche_number = %s,lchg_date = %s,lchg_userid = %s,wx_cond_basic = %s,faa_dist_office = %s,io_last_seen_ntsb = %s
                 WHERE ev_id = %s
                   AND NOT (ntsb_no = %s AND ev_type = %s AND ev_date = %s AND ev_dow = %s AND ev_time = %s AND ev_tmzn = %s AND ev_city = %s AND
                            ev_state = %s AND ev_country = %s AND ev_site_zipcode = %s AND ev_year = %s AND ev_month = %s AND mid_air = %s AND
                            on_ground_collision = %s AND latitude = %s AND longitude = %s AND latlong_acq = %s AND apt_name = %s AND ev_nr_apt_id = %s AND
                            ev_nr_apt_loc = %s AND apt_dist = %s AND apt_dir = %s AND apt_elev = %s AND wx_brief_comp = %s AND wx_src_iic = %s AND
                            wx_obs_time = %s AND wx_obs_dir = %s AND wx_obs_fac_id = %s AND wx_obs_elev = %s AND wx_obs_dist = %s AND wx_obs_tmzn = %s AND
                            light_cond = %s AND sky_cond_nonceil = %s AND sky_nonceil_ht = %s AND sky_ceil_ht = %s AND sky_cond_ceil = %s AND vis_rvr = %s AND
                            vis_rvv = %s AND vis_sm = %s AND wx_temp = %s AND wx_dew_pt = %s AND wind_dir_deg = %s AND wind_dir_ind = %s AND
                            wind_vel_kts = %s AND wind_vel_ind = %s AND gust_ind = %s AND gust_kts = %s AND altimeter = %s AND wx_dens_alt = %s AND
                            wx_int_precip = %s AND metar = %s AND ev_highest_injury = %s AND inj_f_grnd = %s AND inj_m_grnd = %s AND inj_s_grnd = %s AND
                            inj_tot_f = %s AND inj_tot_m = %s AND inj_tot_n = %s AND inj_tot_s = %s AND inj_tot_t = %s AND invest_agy = %s AND
                            ntsb_docket = %s AND ntsb_notf_from = %s AND ntsb_notf_date = %s AND ntsb_notf_tm = %s AND fiche_number = %s AND
                            lchg_date = %s AND lchg_userid = %s AND wx_cond_basic = %s AND faa_dist_office = %s);
                """,
                    (
                        row_mdb.ntsb_no,
                        row_mdb.ev_type,
                        row_mdb.ev_date,
                        row_mdb.ev_dow,
                        row_mdb.ev_time,
                        row_mdb.ev_tmzn,
                        None
                        if row_mdb.ev_city is None
                        else row_mdb.ev_city.rstrip()
                        if len(row_mdb.ev_city.rstrip()) > 0
                        else None,
                        row_mdb.ev_state,
                        row_mdb.ev_country,
                        None
                        if row_mdb.ev_site_zipcode is None
                        else row_mdb.ev_site_zipcode.rstrip()
                        if len(row_mdb.ev_site_zipcode.rstrip()) > 0
                        else None,
                        row_mdb.ev_year,
                        row_mdb.ev_month,
                        row_mdb.mid_air,
                        row_mdb.on_ground_collision,
                        None
                        if row_mdb.latitude is None
                        else row_mdb.latitude.rstrip()
                        if len(row_mdb.latitude.rstrip()) > 0
                        else None,
                        None
                        if row_mdb.longitude is None
                        else row_mdb.longitude.rstrip()
                        if len(row_mdb.longitude.rstrip()) > 0
                        else None,
                        row_mdb.latlong_acq,
                        row_mdb.apt_name,
                        row_mdb.ev_nr_apt_id,
                        row_mdb.ev_nr_apt_loc,
                        row_mdb.apt_dist,
                        row_mdb.apt_dir,
                        row_mdb.apt_elev,
                        row_mdb.wx_brief_comp,
                        row_mdb.wx_src_iic,
                        row_mdb.wx_obs_time,
                        row_mdb.wx_obs_dir,
                        row_mdb.wx_obs_fac_id,
                        row_mdb.wx_obs_elev,
                        row_mdb.wx_obs_dist,
                        row_mdb.wx_obs_tmzn,
                        row_mdb.light_cond,
                        row_mdb.sky_cond_nonceil,
                        row_mdb.sky_nonceil_ht,
                        row_mdb.sky_ceil_ht,
                        row_mdb.sky_cond_ceil,
                        row_mdb.vis_rvr,
                        row_mdb.vis_rvv,
                        row_mdb.vis_sm,
                        row_mdb.wx_temp,
                        row_mdb.wx_dew_pt,
                        row_mdb.wind_dir_deg,
                        row_mdb.wind_dir_ind,
                        row_mdb.wind_vel_kts,
                        row_mdb.wind_vel_ind,
                        row_mdb.gust_ind,
                        row_mdb.gust_kts,
                        row_mdb.altimeter,
                        row_mdb.wx_dens_alt,
                        row_mdb.wx_int_precip,
                        row_mdb.metar,
                        row_mdb.ev_highest_injury,
                        row_mdb.inj_f_grnd,
                        row_mdb.inj_m_grnd,
                        row_mdb.inj_s_grnd,
                        row_mdb.inj_tot_f,
                        row_mdb.inj_tot_m,
                        row_mdb.inj_tot_n,
                        row_mdb.inj_tot_s,
                        row_mdb.inj_tot_t,
                        row_mdb.invest_agy,
                        row_mdb.ntsb_docket,
                        row_mdb.ntsb_notf_from,
                        row_mdb.ntsb_notf_date,
                        row_mdb.ntsb_notf_tm,
                        row_mdb.fiche_number,
                        row_mdb.lchg_date,
                        row_mdb.lchg_userid,
                        row_mdb.wx_cond_basic,
                        row_mdb.faa_dist_office,
                        IO_LAST_SEEN,
                        row_mdb.ev_id,
                        row_mdb.ntsb_no,
                        row_mdb.ev_type,
                        row_mdb.ev_date,
                        row_mdb.ev_dow,
                        row_mdb.ev_time,
                        row_mdb.ev_tmzn,
                        None
                        if row_mdb.ev_city is None
                        else row_mdb.ev_city.rstrip()
                        if len(row_mdb.ev_city.rstrip()) > 0
                        else None,
                        row_mdb.ev_state,
                        row_mdb.ev_country,
                        None
                        if row_mdb.ev_site_zipcode is None
                        else row_mdb.ev_site_zipcode.rstrip()
                        if len(row_mdb.ev_site_zipcode.rstrip()) > 0
                        else None,
                        row_mdb.ev_year,
                        row_mdb.ev_month,
                        row_mdb.mid_air,
                        row_mdb.on_ground_collision,
                        None
                        if row_mdb.latitude is None
                        else row_mdb.latitude.rstrip()
                        if len(row_mdb.latitude.rstrip()) > 0
                        else None,
                        None
                        if row_mdb.longitude is None
                        else row_mdb.longitude.rstrip()
                        if len(row_mdb.longitude.rstrip()) > 0
                        else None,
                        row_mdb.latlong_acq,
                        row_mdb.apt_name,
                        row_mdb.ev_nr_apt_id,
                        row_mdb.ev_nr_apt_loc,
                        row_mdb.apt_dist,
                        row_mdb.apt_dir,
                        row_mdb.apt_elev,
                        row_mdb.wx_brief_comp,
                        row_mdb.wx_src_iic,
                        row_mdb.wx_obs_time,
                        row_mdb.wx_obs_dir,
                        row_mdb.wx_obs_fac_id,
                        row_mdb.wx_obs_elev,
                        row_mdb.wx_obs_dist,
                        row_mdb.wx_obs_tmzn,
                        row_mdb.light_cond,
                        row_mdb.sky_cond_nonceil,
                        row_mdb.sky_nonceil_ht,
                        row_mdb.sky_ceil_ht,
                        row_mdb.sky_cond_ceil,
                        row_mdb.vis_rvr,
                        row_mdb.vis_rvv,
                        row_mdb.vis_sm,
                        row_mdb.wx_temp,
                        row_mdb.wx_dew_pt,
                        row_mdb.wind_dir_deg,
                        row_mdb.wind_dir_ind,
                        row_mdb.wind_vel_kts,
                        row_mdb.wind_vel_ind,
                        row_mdb.gust_ind,
                        row_mdb.gust_kts,
                        row_mdb.altimeter,
                        row_mdb.wx_dens_alt,
                        row_mdb.wx_int_precip,
                        row_mdb.metar,
                        row_mdb.ev_highest_injury,
                        row_mdb.inj_f_grnd,
                        row_mdb.inj_m_grnd,
                        row_mdb.inj_s_grnd,
                        row_mdb.inj_tot_f,
                        row_mdb.inj_tot_m,
                        row_mdb.inj_tot_n,
                        row_mdb.inj_tot_s,
                        row_mdb.inj_tot_t,
                        row_mdb.invest_agy,
                        row_mdb.ntsb_docket,
                        row_mdb.ntsb_notf_from,
                        row_mdb.ntsb_notf_date,
                        row_mdb.ntsb_notf_tm,
                        row_mdb.fiche_number,
                        row_mdb.lchg_date,
                        row_mdb.lchg_userid,
                        row_mdb.wx_cond_basic,
                        row_mdb.faa_dist_office,
                    ),
                )
            else:
                # pylint: disable=line-too-long
                cur_pg.execute(
                    """
                UPDATE events
                   SET ntsb_no = %s,ev_type = %s,ev_date = %s,ev_dow = %s,ev_time = %s,ev_tmzn = %s,ev_city = %s,ev_state = %s,ev_country = %s,
                   ev_site_zipcode = %s,ev_year = %s,ev_month = %s,mid_air = %s,on_ground_collision = %s,latitude = %s,longitude = %s,latlong_acq = %s,
                   apt_name = %s,ev_nr_apt_id = %s,ev_nr_apt_loc = %s,apt_dist = %s,apt_dir = %s,apt_elev = %s,wx_brief_comp = %s,wx_src_iic = %s,
                   wx_obs_time = %s,wx_obs_dir = %s,wx_obs_fac_id = %s,wx_obs_elev = %s,wx_obs_dist = %s,wx_obs_tmzn = %s,light_cond = %s,
                   sky_cond_nonceil = %s,sky_nonceil_ht = %s,sky_ceil_ht = %s,sky_cond_ceil = %s,vis_rvr = %s,vis_rvv = %s,vis_sm = %s,
                   wx_temp = %s,wx_dew_pt = %s,wind_dir_deg = %s,wind_dir_ind = %s,wind_vel_kts = %s,wind_vel_ind = %s,gust_ind = %s,gust_kts = %s,
                   altimeter = %s,wx_dens_alt = %s,wx_int_precip = %s,metar = %s,ev_highest_injury = %s,inj_f_grnd = %s,inj_m_grnd = %s,inj_s_grnd = %s,
                   inj_tot_f = %s,inj_tot_m = %s,inj_tot_n = %s,inj_tot_s = %s,inj_tot_t = %s,invest_agy = %s,ntsb_docket = %s,ntsb_notf_from = %s,
                   ntsb_notf_date = %s,ntsb_notf_tm = %s,fiche_number = %s,lchg_date = %s,lchg_userid = %s,wx_cond_basic = %s,faa_dist_office = %s,
                   dec_latitude = %s,dec_longitude = %s,io_last_seen_ntsb = %s
                 WHERE ev_id = %s
                   AND NOT (ntsb_no = %s AND ev_type = %s AND ev_date = %s AND ev_dow = %s AND ev_time = %s AND ev_tmzn = %s AND ev_city = %s AND
                            ev_state = %s AND ev_country = %s AND ev_site_zipcode = %s AND ev_year = %s AND ev_month = %s AND mid_air = %s AND
                            on_ground_collision = %s AND latitude = %s AND longitude = %s AND latlong_acq = %s AND apt_name = %s AND ev_nr_apt_id = %s AND
                            ev_nr_apt_loc = %s AND apt_dist = %s AND apt_dir = %s AND apt_elev = %s AND wx_brief_comp = %s AND wx_src_iic = %s AND
                            wx_obs_time = %s AND wx_obs_dir = %s AND wx_obs_fac_id = %s AND wx_obs_elev = %s AND wx_obs_dist = %s AND wx_obs_tmzn = %s AND
                            light_cond = %s AND sky_cond_nonceil = %s AND sky_nonceil_ht = %s AND sky_ceil_ht = %s AND sky_cond_ceil = %s AND
                            vis_rvr = %s AND vis_rvv = %s AND vis_sm = %s AND wx_temp = %s AND wx_dew_pt = %s AND wind_dir_deg = %s AND
                            wind_dir_ind = %s AND wind_vel_kts = %s AND wind_vel_ind = %s AND gust_ind = %s AND gust_kts = %s AND altimeter = %s AND
                            wx_dens_alt = %s AND wx_int_precip = %s AND metar = %s AND ev_highest_injury = %s AND inj_f_grnd = %s AND inj_m_grnd = %s AND
                            inj_s_grnd = %s AND inj_tot_f = %s AND inj_tot_m = %s AND inj_tot_n = %s AND inj_tot_s = %s AND inj_tot_t = %s AND
                            invest_agy = %s AND ntsb_docket = %s AND ntsb_notf_from = %s AND ntsb_notf_date = %s AND ntsb_notf_tm = %s AND
                            fiche_number = %s AND lchg_date = %s AND lchg_userid = %s AND wx_cond_basic = %s AND faa_dist_office = %s
                            AND dec_latitude = %s AND dec_longitude = %s);
                """,
                    (
                        row_mdb.ntsb_no,
                        row_mdb.ev_type,
                        row_mdb.ev_date,
                        row_mdb.ev_dow,
                        row_mdb.ev_time,
                        row_mdb.ev_tmzn,
                        None
                        if row_mdb.ev_city is None
                        else row_mdb.ev_city.rstrip()
                        if len(row_mdb.ev_city.rstrip()) > 0
                        else None,
                        row_mdb.ev_state,
                        row_mdb.ev_country,
                        None
                        if row_mdb.ev_site_zipcode is None
                        else row_mdb.ev_site_zipcode.rstrip()
                        if len(row_mdb.ev_site_zipcode.rstrip()) > 0
                        else None,
                        row_mdb.ev_year,
                        row_mdb.ev_month,
                        row_mdb.mid_air,
                        row_mdb.on_ground_collision,
                        None
                        if row_mdb.latitude is None
                        else row_mdb.latitude.rstrip()
                        if len(row_mdb.latitude.rstrip()) > 0
                        else None,
                        None
                        if row_mdb.longitude is None
                        else row_mdb.longitude.rstrip()
                        if len(row_mdb.longitude.rstrip()) > 0
                        else None,
                        row_mdb.latlong_acq,
                        row_mdb.apt_name,
                        row_mdb.ev_nr_apt_id,
                        row_mdb.ev_nr_apt_loc,
                        row_mdb.apt_dist,
                        row_mdb.apt_dir,
                        row_mdb.apt_elev,
                        row_mdb.wx_brief_comp,
                        row_mdb.wx_src_iic,
                        row_mdb.wx_obs_time,
                        row_mdb.wx_obs_dir,
                        row_mdb.wx_obs_fac_id,
                        row_mdb.wx_obs_elev,
                        row_mdb.wx_obs_dist,
                        row_mdb.wx_obs_tmzn,
                        row_mdb.light_cond,
                        row_mdb.sky_cond_nonceil,
                        row_mdb.sky_nonceil_ht,
                        row_mdb.sky_ceil_ht,
                        row_mdb.sky_cond_ceil,
                        row_mdb.vis_rvr,
                        row_mdb.vis_rvv,
                        row_mdb.vis_sm,
                        row_mdb.wx_temp,
                        row_mdb.wx_dew_pt,
                        row_mdb.wind_dir_deg,
                        row_mdb.wind_dir_ind,
                        row_mdb.wind_vel_kts,
                        row_mdb.wind_vel_ind,
                        row_mdb.gust_ind,
                        row_mdb.gust_kts,
                        row_mdb.altimeter,
                        row_mdb.wx_dens_alt,
                        row_mdb.wx_int_precip,
                        row_mdb.metar,
                        row_mdb.ev_highest_injury,
                        row_mdb.inj_f_grnd,
                        row_mdb.inj_m_grnd,
                        row_mdb.inj_s_grnd,
                        row_mdb.inj_tot_f,
                        row_mdb.inj_tot_m,
                        row_mdb.inj_tot_n,
                        row_mdb.inj_tot_s,
                        row_mdb.inj_tot_t,
                        row_mdb.invest_agy,
                        row_mdb.ntsb_docket,
                        row_mdb.ntsb_notf_from,
                        row_mdb.ntsb_notf_date,
                        row_mdb.ntsb_notf_tm,
                        row_mdb.fiche_number,
                        row_mdb.lchg_date,
                        row_mdb.lchg_userid,
                        row_mdb.wx_cond_basic,
                        row_mdb.faa_dist_office,
                        row_mdb.dec_latitude,
                        row_mdb.dec_longitude,
                        IO_LAST_SEEN,
                        row_mdb.ev_id,
                        row_mdb.ntsb_no,
                        row_mdb.ev_type,
                        row_mdb.ev_date,
                        row_mdb.ev_dow,
                        row_mdb.ev_time,
                        row_mdb.ev_tmzn,
                        None
                        if row_mdb.ev_city is None
                        else row_mdb.ev_city.rstrip()
                        if len(row_mdb.ev_city.rstrip()) > 0
                        else None,
                        row_mdb.ev_state,
                        row_mdb.ev_country,
                        None
                        if row_mdb.ev_site_zipcode is None
                        else row_mdb.ev_site_zipcode.rstrip()
                        if len(row_mdb.ev_site_zipcode.rstrip()) > 0
                        else None,
                        row_mdb.ev_year,
                        row_mdb.ev_month,
                        row_mdb.mid_air,
                        row_mdb.on_ground_collision,
                        None
                        if row_mdb.latitude is None
                        else row_mdb.latitude.rstrip()
                        if len(row_mdb.latitude.rstrip()) > 0
                        else None,
                        None
                        if row_mdb.longitude is None
                        else row_mdb.longitude.rstrip()
                        if len(row_mdb.longitude.rstrip()) > 0
                        else None,
                        row_mdb.latlong_acq,
                        row_mdb.apt_name,
                        row_mdb.ev_nr_apt_id,
                        row_mdb.ev_nr_apt_loc,
                        row_mdb.apt_dist,
                        row_mdb.apt_dir,
                        row_mdb.apt_elev,
                        row_mdb.wx_brief_comp,
                        row_mdb.wx_src_iic,
                        row_mdb.wx_obs_time,
                        row_mdb.wx_obs_dir,
                        row_mdb.wx_obs_fac_id,
                        row_mdb.wx_obs_elev,
                        row_mdb.wx_obs_dist,
                        row_mdb.wx_obs_tmzn,
                        row_mdb.light_cond,
                        row_mdb.sky_cond_nonceil,
                        row_mdb.sky_nonceil_ht,
                        row_mdb.sky_ceil_ht,
                        row_mdb.sky_cond_ceil,
                        row_mdb.vis_rvr,
                        row_mdb.vis_rvv,
                        row_mdb.vis_sm,
                        row_mdb.wx_temp,
                        row_mdb.wx_dew_pt,
                        row_mdb.wind_dir_deg,
                        row_mdb.wind_dir_ind,
                        row_mdb.wind_vel_kts,
                        row_mdb.wind_vel_ind,
                        row_mdb.gust_ind,
                        row_mdb.gust_kts,
                        row_mdb.altimeter,
                        row_mdb.wx_dens_alt,
                        row_mdb.wx_int_precip,
                        row_mdb.metar,
                        row_mdb.ev_highest_injury,
                        row_mdb.inj_f_grnd,
                        row_mdb.inj_m_grnd,
                        row_mdb.inj_s_grnd,
                        row_mdb.inj_tot_f,
                        row_mdb.inj_tot_m,
                        row_mdb.inj_tot_n,
                        row_mdb.inj_tot_s,
                        row_mdb.inj_tot_t,
                        row_mdb.invest_agy,
                        row_mdb.ntsb_docket,
                        row_mdb.ntsb_notf_from,
                        row_mdb.ntsb_notf_date,
                        row_mdb.ntsb_notf_tm,
                        row_mdb.fiche_number,
                        row_mdb.lchg_date,
                        row_mdb.lchg_userid,
                        row_mdb.wx_cond_basic,
                        row_mdb.faa_dist_office,
                        row_mdb.dec_latitude,
                        row_mdb.dec_longitude,
                    ),
                )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"ev_year={row_mdb.ev_year}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table events_sequence.
# ------------------------------------------------------------------
def _load_table_events_sequence(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table events_sequence."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO events_sequence (
                   ev_id,aircraft_key,occurrence_no,occurrence_code,occurrence_description,phase_no,eventsoe_no,defining_ev,lchg_date,lchg_userid,
                   io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.Occurrence_No,
                    row_mdb.Occurrence_Code,
                    row_mdb.Occurrence_Description,
                    row_mdb.phase_no,
                    row_mdb.eventsoe_no,
                    row_mdb.Defining_ev,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE events_sequence
               SET occurrence_code = %s,occurrence_description = %s,phase_no = %s,eventsoe_no = %s,defining_ev = %s,lchg_date = %s,lchg_userid = %s,
                   io_last_seen_ntsb = %s
             WHERE occurrence_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (occurrence_code = %s AND occurrence_description = %s AND phase_no = %s AND eventsoe_no = %s AND defining_ev = %s AND
                        lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.Occurrence_Code,
                    row_mdb.Occurrence_Description,
                    row_mdb.phase_no,
                    row_mdb.eventsoe_no,
                    row_mdb.Defining_ev,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.Occurrence_No,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.Occurrence_Code,
                    row_mdb.Occurrence_Description,
                    row_mdb.phase_no,
                    row_mdb.eventsoe_no,
                    row_mdb.Defining_ev,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"Aircraft_Key={row_mdb.Aircraft_Key} "
                        + f"Occurrence_No={row_mdb.Occurrence_No}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table findings.
# ------------------------------------------------------------------
def _load_table_findings(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table findings."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO findings (
                   ev_id,aircraft_key,finding_no,finding_code,finding_description,category_no,subcategory_no,section_no,subsection_no,modifier_no,
                   cause_factor,lchg_date,lchg_userid,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.finding_no,
                    row_mdb.finding_code,
                    row_mdb.finding_description,
                    row_mdb.category_no,
                    row_mdb.subcategory_no,
                    row_mdb.section_no,
                    row_mdb.subsection_no,
                    row_mdb.modifier_no,
                    row_mdb.Cause_Factor,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE findings
               SET finding_code = %s,finding_description = %s,category_no = %s,subcategory_no = %s,section_no = %s,subsection_no = %s,modifier_no = %s,
                   cause_factor = %s,lchg_date = %s,lchg_userid = %s,io_last_seen_ntsb = %s
             WHERE finding_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (finding_code = %s AND finding_description = %s AND category_no = %s AND subcategory_no = %s AND section_no = %s AND
                        subsection_no = %s AND modifier_no = %s AND cause_factor = %s AND lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.finding_code,
                    row_mdb.finding_description,
                    row_mdb.category_no,
                    row_mdb.subcategory_no,
                    row_mdb.section_no,
                    row_mdb.subsection_no,
                    row_mdb.modifier_no,
                    row_mdb.Cause_Factor,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.finding_no,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.finding_code,
                    row_mdb.finding_description,
                    row_mdb.category_no,
                    row_mdb.subcategory_no,
                    row_mdb.section_no,
                    row_mdb.subsection_no,
                    row_mdb.modifier_no,
                    row_mdb.Cause_Factor,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"finding_no={row_mdb.finding_no}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table flight_crew.
# ------------------------------------------------------------------
def _load_table_flight_crew(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table flight_crew."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO flight_crew (
                   ev_id,aircraft_key,crew_no,crew_category,crew_age,crew_sex,crew_city,crew_res_state,crew_res_country,med_certf,med_crtf_vldty,
                   date_lst_med,crew_rat_endorse,crew_inj_level,seatbelts_used,shldr_harn_used,crew_tox_perf,seat_occ_pic,pc_profession,bfr,bfr_date,
                   ft_as_of,lchg_date,lchg_userid,seat_occ_row,infl_rest_inst,infl_rest_depl,child_restraint,med_crtf_limit,mr_faa_med_certf,pilot_flying,
                   available_restraint,restraint_used,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.crew_no,
                    row_mdb.crew_category,
                    row_mdb.crew_age,
                    row_mdb.crew_sex,
                    row_mdb.crew_city,
                    row_mdb.crew_res_state,
                    row_mdb.crew_res_country,
                    row_mdb.med_certf,
                    row_mdb.med_crtf_vldty,
                    row_mdb.date_lst_med,
                    row_mdb.crew_rat_endorse,
                    row_mdb.crew_inj_level,
                    row_mdb.seatbelts_used,
                    row_mdb.shldr_harn_used,
                    row_mdb.crew_tox_perf,
                    row_mdb.seat_occ_pic,
                    row_mdb.pc_profession,
                    row_mdb.bfr,
                    row_mdb.bfr_date,
                    row_mdb.ft_as_of,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    row_mdb.seat_occ_row,
                    row_mdb.infl_rest_inst,
                    row_mdb.infl_rest_depl,
                    row_mdb.child_restraint,
                    row_mdb.med_crtf_limit,
                    row_mdb.mr_faa_med_certf,
                    row_mdb.pilot_flying,
                    row_mdb.available_restraint,
                    row_mdb.restraint_used,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE flight_crew
               SET crew_category = %s,crew_age = %s,crew_sex = %s,crew_city = %s,crew_res_state = %s,crew_res_country = %s,med_certf = %s,
                   med_crtf_vldty = %s,date_lst_med = %s,crew_rat_endorse = %s,crew_inj_level = %s,seatbelts_used = %s,shldr_harn_used = %s,
                   crew_tox_perf = %s,seat_occ_pic = %s,pc_profession = %s,bfr = %s,bfr_date = %s,ft_as_of = %s,lchg_date = %s,lchg_userid = %s,
                   seat_occ_row = %s,infl_rest_inst = %s,infl_rest_depl = %s,child_restraint = %s,med_crtf_limit = %s,mr_faa_med_certf = %s,
                   pilot_flying = %s,available_restraint = %s,restraint_used = %s,io_last_seen_ntsb = %s
             WHERE crew_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (crew_category = %s AND crew_age = %s AND crew_sex = %s AND crew_city = %s AND crew_res_state = %s AND crew_res_country = %s AND
                        med_certf = %s AND med_crtf_vldty = %s AND date_lst_med = %s AND crew_rat_endorse = %s AND crew_inj_level = %s AND
                        seatbelts_used = %s AND shldr_harn_used = %s AND crew_tox_perf = %s AND seat_occ_pic = %s AND pc_profession = %s AND bfr = %s AND
                        bfr_date = %s AND ft_as_of = %s AND lchg_date = %s AND lchg_userid = %s AND seat_occ_row = %s AND infl_rest_inst = %s AND
                        infl_rest_depl = %s AND child_restraint = %s AND med_crtf_limit = %s AND mr_faa_med_certf = %s AND pilot_flying = %s AND
                        available_restraint = %s AND restraint_used = %s);
            """,
                (
                    row_mdb.crew_category,
                    row_mdb.crew_age,
                    row_mdb.crew_sex,
                    row_mdb.crew_city,
                    row_mdb.crew_res_state,
                    row_mdb.crew_res_country,
                    row_mdb.med_certf,
                    row_mdb.med_crtf_vldty,
                    row_mdb.date_lst_med,
                    row_mdb.crew_rat_endorse,
                    row_mdb.crew_inj_level,
                    row_mdb.seatbelts_used,
                    row_mdb.shldr_harn_used,
                    row_mdb.crew_tox_perf,
                    row_mdb.seat_occ_pic,
                    row_mdb.pc_profession,
                    row_mdb.bfr,
                    row_mdb.bfr_date,
                    row_mdb.ft_as_of,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    row_mdb.seat_occ_row,
                    row_mdb.infl_rest_inst,
                    row_mdb.infl_rest_depl,
                    row_mdb.child_restraint,
                    row_mdb.med_crtf_limit,
                    row_mdb.mr_faa_med_certf,
                    row_mdb.pilot_flying,
                    row_mdb.available_restraint,
                    row_mdb.restraint_used,
                    IO_LAST_SEEN,
                    row_mdb.crew_no,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.crew_category,
                    row_mdb.crew_age,
                    row_mdb.crew_sex,
                    row_mdb.crew_city,
                    row_mdb.crew_res_state,
                    row_mdb.crew_res_country,
                    row_mdb.med_certf,
                    row_mdb.med_crtf_vldty,
                    row_mdb.date_lst_med,
                    row_mdb.crew_rat_endorse,
                    row_mdb.crew_inj_level,
                    row_mdb.seatbelts_used,
                    row_mdb.shldr_harn_used,
                    row_mdb.crew_tox_perf,
                    row_mdb.seat_occ_pic,
                    row_mdb.pc_profession,
                    row_mdb.bfr,
                    row_mdb.bfr_date,
                    row_mdb.ft_as_of,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    row_mdb.seat_occ_row,
                    row_mdb.infl_rest_inst,
                    row_mdb.infl_rest_depl,
                    row_mdb.child_restraint,
                    row_mdb.med_crtf_limit,
                    row_mdb.mr_faa_med_certf,
                    row_mdb.pilot_flying,
                    row_mdb.available_restraint,
                    row_mdb.restraint_used,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"crew_no={row_mdb.crew_no}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table flight_time.
# ------------------------------------------------------------------
def _load_table_flight_time(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table flight_time."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO flight_time (
                   ev_id,aircraft_key,crew_no,flight_type,flight_craft,flight_hours,lchg_date,lchg_userid,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.crew_no,
                    row_mdb.flight_type,
                    row_mdb.flight_craft,
                    row_mdb.flight_hours,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE flight_time
               SET flight_hours = %s,lchg_date = %s,lchg_userid = %s,io_last_seen_ntsb = %s
             WHERE flight_craft = %s AND flight_type = %s AND crew_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (flight_hours = %s AND lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.flight_hours,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.flight_craft,
                    row_mdb.flight_type,
                    row_mdb.crew_no,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.flight_hours,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"crew_no={row_mdb.crew_no} "
                        + f"flight_type={row_mdb.flight_type} "
                        + f"flight_craft={row_mdb.flight_craft}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table injury.
# ------------------------------------------------------------------
def _load_table_injury(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table injury."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO injury (
                   ev_id,aircraft_key,inj_person_category,injury_level,inj_person_count,lchg_date,lchg_userid,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.inj_person_category,
                    row_mdb.injury_level,
                    row_mdb.inj_person_count,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE injury
               SET inj_person_count = %s,lchg_date = %s,lchg_userid = %s,io_last_seen_ntsb = %s
             WHERE injury_level = %s AND inj_person_category = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (inj_person_count = %s AND lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.inj_person_count,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.injury_level,
                    row_mdb.inj_person_category,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.inj_person_count,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"inj_person_category={row_mdb.inj_person_category} "
                        + f"injury_level={row_mdb.injury_level}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Determine and load city averages.
# ------------------------------------------------------------------
def _load_table_io_lat_lng_average(conn_pg, cur_pg) -> None:
    """Determine and load city averages."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Delete averaged data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_lat_lng
     WHERE
           source = %s;
    """,
        (io_glob.SOURCE_AVERAGE,),
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()

        # INFO.00.063 Processed data source '{data_source}'
        io_utils.progress_msg(
            io_glob.INFO_00_063.replace("{data_source}", io_glob.SOURCE_AVERAGE)
        )
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")
        io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Insert averaged data.
    # ------------------------------------------------------------------
    # INFO.00.062 Database table io_lat_lng: Load the averaged city data
    io_utils.progress_msg(io_glob.INFO_00_062)
    io_utils.progress_msg("-" * 80)

    count_duplicates = 0
    count_insert = 0
    count_select = 0

    conn_pg_2, cur_pg_2 = db_utils.get_postgres_cursor()

    conn_pg_2.set_session(autocommit=False)

    # pylint: disable=line-too-long
    cur_pg_2.execute(
        f"""
    SELECT country, state, city, sum(dec_latitude)/count(*) dec_latitude, sum(dec_longitude)/count(*) dec_longitude
      FROM io_lat_lng
     WHERE type = '{io_glob.IO_LAT_LNG_TYPE_ZIPCODE}'
     GROUP BY country, state, city;
        """,
    )

    for row_pg in cur_pg_2.fetchall():
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            cur_pg.execute(
                """
            INSERT INTO io_lat_lng (
                   type,
                   country,
                   state,
                   city,
                   dec_latitude,
                   dec_longitude,
                   source,
                   first_processed
                   ) VALUES (
                   %s,%s,%s,%s,%s,
                   %s,%s,%s
                   )
            ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
            DO NOTHING;
            """,
                (
                    io_glob.IO_LAT_LNG_TYPE_CITY,
                    row_pg["country"],  # type: ignore
                    row_pg["state"],  # type: ignore
                    row_pg["city"],  # type: ignore
                    row_pg["dec_latitude"],  # type: ignore
                    row_pg["dec_longitude"],  # type: ignore
                    io_glob.SOURCE_AVERAGE,
                    datetime.now(),
                ),
            )
            count_insert += 1
        except UniqueViolation:
            count_duplicates += cur_pg.rowcount

    conn_pg.commit()

    cur_pg_2.close()
    conn_pg_2.close()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    if count_duplicates > 0:
        io_utils.progress_msg(f"Number rows duplicate: {str(count_duplicates):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table narratives.
# ------------------------------------------------------------------
def _load_table_narratives(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table narratives."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            cur_pg.execute(
                # pylint: disable=line-too-long
                """
            INSERT INTO narratives (
                   ev_id,aircraft_key,narr_accp,narr_accf,narr_cause,narr_inc,lchg_date,lchg_userid,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.narr_accp,
                    row_mdb.narr_accf,
                    row_mdb.narr_cause,
                    row_mdb.narr_inc,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            UPDATE narratives
               SET narr_accp = %s,narr_accf = %s,narr_cause = %s,narr_inc = %s,lchg_date = %s,lchg_userid = %s,io_last_seen_ntsb = %s
             WHERE aircraft_key = %s AND ev_id = %s
               AND NOT (narr_accp = %s AND narr_accf = %s AND narr_cause = %s AND narr_inc = %s AND lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.narr_accp,
                    row_mdb.narr_accf,
                    row_mdb.narr_cause,
                    row_mdb.narr_inc,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.narr_accp,
                    row_mdb.narr_accf,
                    row_mdb.narr_cause,
                    row_mdb.narr_inc,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key}"
                    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table ntsb_admin.
# ------------------------------------------------------------------
def _load_table_ntsb_admin(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table ntsb_admin."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        if row_mdb.ev_id in ["20210527103155", "20220328104839"]:
            # ERROR_00_946 = "ERROR.00.946 The ev_id '{ev_id}' is missing in database table events"
            io_utils.progress_msg(
                io_glob.ERROR_00_946.replace("{ev_id}", row_mdb.ev_id)
            )
            continue

        try:
            cur_pg.execute(
                """
            INSERT INTO ntsb_admin (
                   ev_id,rec_stat,approval_date,lchg_userid,lchg_date,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.rec_stat,
                    row_mdb.approval_date,
                    row_mdb.lchg_userid,
                    row_mdb.lchg_date,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            UPDATE ntsb_admin
               SET rec_stat = %s,approval_date = %s,lchg_userid = %s,lchg_date = %s,io_last_seen_ntsb = %s
             WHERE ev_id = %s
               AND NOT (rec_stat = %s AND approval_date = %s AND lchg_userid = %s AND lchg_date = %s);
            """,
                (
                    row_mdb.rec_stat,
                    row_mdb.approval_date,
                    row_mdb.lchg_userid,
                    row_mdb.lchg_date,
                    IO_LAST_SEEN,
                    row_mdb.ev_id,
                    row_mdb.rec_stat,
                    row_mdb.approval_date,
                    row_mdb.lchg_userid,
                    row_mdb.lchg_date,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(f"Updated  ev_id={row_mdb.ev_id}")
        except ForeignKeyViolation:
            # ERROR_00_947 = "ERROR.00.947 The ev_id '{ev_id}' is missing in database table events"
            io_utils.progress_msg(
                io_glob.ERROR_00_947.replace("{ev_id}", row_mdb.ev_id)
            )
            conn_pg.rollback()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table occurrences.
# ------------------------------------------------------------------
def _load_table_occurrences(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table occurrences."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        if row_mdb.ev_id in [
            "20001213X25705",
            "20001213X29793",
            "20001213X30454",
            "20001213X31758",
            "20001213X32752",
            "20001218X45444",
            "20001218X45446",
        ]:
            # ERROR_00_946 = "ERROR.00.946 The ev_id '{ev_id}' is missing in database table events"
            io_utils.progress_msg(
                io_glob.ERROR_00_946.replace("{ev_id}", row_mdb.ev_id)
            )
            continue

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO occurrences (
                   ev_id,aircraft_key,occurrence_no,occurrence_code,phase_of_flight,altitude,lchg_date,lchg_userid,io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.Occurrence_No,
                    row_mdb.Occurrence_Code,
                    row_mdb.Phase_of_Flight,
                    row_mdb.Altitude,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE occurrences
               SET occurrence_code = %s,phase_of_flight = %s,altitude = %s,lchg_date = %s,lchg_userid = %s,io_last_seen_ntsb = %s
             WHERE occurrence_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (occurrence_code = %s AND phase_of_flight = %s AND altitude = %s AND lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.Occurrence_Code,
                    row_mdb.Phase_of_Flight,
                    row_mdb.Altitude,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.Occurrence_No,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.Occurrence_Code,
                    row_mdb.Phase_of_Flight,
                    row_mdb.Altitude,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"Occurrence_No={row_mdb.Occurrence_No}"
                    )
        except ForeignKeyViolation:
            # ERROR_00_947 = "ERROR.00.947 The ev_id '{ev_id}' is missing in database table events"
            io_utils.progress_msg(
                io_glob.ERROR_00_947.replace("{ev_id}", row_mdb.ev_id)
            )
            conn_pg.rollback()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the data from database table seq_of_events.
# ------------------------------------------------------------------
def _load_table_seq_of_events(
    msaccess: str,
    table_name: str,
    cur_mdb: pyodbc.Cursor,
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    """Load the data from database table seq_of_events."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    count_insert = 0
    count_select = 0
    count_update = 0

    io_utils.progress_msg("")
    io_utils.progress_msg(
        f"Database table       : {table_name.lower():30}" + "<" + "-" * 35
    )

    cur_mdb.execute(f"SELECT * FROM {table_name};")

    rows_mdb = cur_mdb.fetchall()

    for row_mdb in rows_mdb:
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        if row_mdb.ev_id in [
            "20020917X03487",
        ]:
            # ERROR_00_946 = "ERROR.00.946 The ev_id '{ev_id}' is missing in database table events"
            io_utils.progress_msg(
                io_glob.ERROR_00_946.replace("{ev_id}", row_mdb.ev_id)
            )
            continue

        try:
            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO seq_of_events (
                   ev_id,aircraft_key,occurrence_no,seq_event_no,group_code,subj_code,cause_factor,modifier_code,person_code,lchg_date,lchg_userid,
                   io_last_seen_ntsb
                   ) VALUES (
                   %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
            """,
                (
                    row_mdb.ev_id,
                    row_mdb.Aircraft_Key,
                    row_mdb.Occurrence_No,
                    row_mdb.seq_event_no,
                    row_mdb.group_code,
                    row_mdb.Subj_Code,
                    row_mdb.Cause_Factor,
                    row_mdb.Modifier_Code,
                    row_mdb.Person_Code,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                ),
            )
            count_insert += 1
        except UniqueViolation:
            cur_pg.execute(
                """
            UPDATE seq_of_events
               SET aircraft_key = %s,occurrence_no = %s,seq_event_no = %s,group_code = %s,subj_code = %s,cause_factor = %s,modifier_code = %s,
                   person_code = %s,lchg_date = %s,lchg_userid = %s,io_last_seen_ntsb = %s
             WHERE group_code = %s AND seq_event_no = %s AND occurrence_no = %s AND aircraft_key = %s AND ev_id = %s
               AND NOT (subj_code = %s AND cause_factor = %s AND modifier_code = %s AND person_code = %s AND lchg_date = %s AND lchg_userid = %s);
            """,
                (
                    row_mdb.Aircraft_Key,
                    row_mdb.Occurrence_No,
                    row_mdb.seq_event_no,
                    row_mdb.group_code,
                    row_mdb.Subj_Code,
                    row_mdb.Cause_Factor,
                    row_mdb.Modifier_Code,
                    row_mdb.Person_Code,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                    IO_LAST_SEEN,
                    row_mdb.group_code,
                    row_mdb.seq_event_no,
                    row_mdb.Occurrence_No,
                    row_mdb.Aircraft_Key,
                    row_mdb.ev_id,
                    row_mdb.Subj_Code,
                    row_mdb.Cause_Factor,
                    row_mdb.Modifier_Code,
                    row_mdb.Person_Code,
                    row_mdb.lchg_date,
                    row_mdb.lchg_userid,
                ),
            )
            if cur_pg.rowcount > 0:
                count_update += cur_pg.rowcount
                if msaccess not in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
                    io_utils.progress_msg(
                        f"Updated  ev_id={row_mdb.ev_id} "
                        + f"aircraft_key={row_mdb.Aircraft_Key} "
                        + f"Occurrence_No={row_mdb.Occurrence_No} "
                        + f"seq_event_no={row_mdb.seq_event_no} "
                        + f"group_code={row_mdb.group_code}"
                    )
        except ForeignKeyViolation:
            # ERROR_00_947 = "ERROR.00.947 The ev_id '{ev_id}' is missing in database table events"
            io_utils.progress_msg(
                io_glob.ERROR_00_947.replace("{ev_id}", row_mdb.ev_id)
            )
            conn_pg.rollback()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")
    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load ZIP Code Database data.
# ------------------------------------------------------------------
def _load_zip_codes_org_data() -> None:
    """Load ZIP Code Database data."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------
    filename_xlsx = io_config.settings.download_file_zip_codes_org_xlsx

    if not os.path.isfile(filename_xlsx):
        # ERROR.00.935 The Zip Code Database file '{filename}' is missing
        io_utils.terminate_fatal(
            io_glob.ERROR_00_935.replace("{filename}", filename_xlsx)
        )

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Delete existing data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_lat_lng
     WHERE SOURCE = %s;
    """,
        (io_glob.SOURCE_ZCO_ZIP_CODES,),
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()

        # INFO.00.063 Processed data source '{data_source}'
        io_utils.progress_msg(
            io_glob.INFO_00_063.replace("{data_source}", io_glob.SOURCE_ZCO_ZIP_CODES)
        )
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")
        io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Insert new data.
    # ------------------------------------------------------------------
    _load_zip_codes_org_data_zips(conn_pg, cur_pg, filename_xlsx)
    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Delete and insert averaged data.
    # ------------------------------------------------------------------
    _load_table_io_lat_lng_average(conn_pg, cur_pg)

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------
    db_utils.upd_io_processed_files(
        io_config.settings.download_file_zip_codes_org_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the estimated zip code data from the ZIP Code Database
# into the PostgreSQL database.
# ------------------------------------------------------------------
def _load_zip_codes_org_data_zips(conn_pg, cur_pg, filename) -> None:
    """Load the estimated zip code data from the ZIP Code Database into the
    PostgreSQL database."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    # INFO.00.061 Database table io_lat_lng: Load the estimated zip code data
    io_utils.progress_msg(io_glob.INFO_00_061)
    io_utils.progress_msg("-" * 80)

    count_upsert = 0
    count_select = 0

    zip_idx = 0
    type_idx = 1
    primary_city_idx = 3
    acceptable_cities_idx = 4
    state_idx = 6
    country_idx = 11
    latitude_idx = 12
    longitude_idx = 13

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        zipcode = f"{row[zip_idx].value:05}".rstrip()
        if zipcode == "zip":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        if row[type_idx].value != "STANDARD" or row[country_idx].value != "US":
            continue

        primary_city = [row[primary_city_idx].value.upper().rstrip()]
        acceptable_cities = (
            row[acceptable_cities_idx].value.upper().split(",")
            if row[acceptable_cities_idx].value
            else []
        )
        state = row[state_idx].value.upper().rstrip()
        lat = row[latitude_idx].value
        lng = row[longitude_idx].value

        for city in acceptable_cities:
            primary_city.append(city.rstrip())

        for city in primary_city:
            if city and city.rstrip():
                # pylint: disable=line-too-long
                cur_pg.execute(
                    """
                INSERT INTO io_lat_lng AS ill (
                       type,
                       country,
                       state,
                       city,
                       zipcode,
                       dec_latitude,
                       dec_longitude,
                       source,
                       first_processed
                       ) VALUES (
                       %s,%s,%s,%s,%s,
                       %s,%s,%s,%s
                       )
                ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
                DO UPDATE
                   SET dec_latitude = %s,
                       dec_longitude = %s,
                       source = %s,
                       last_processed = %s
                 WHERE ill.type = %s
                   AND ill.country = %s
                   AND ill.state = %s
                   AND ill.city = %s
                   AND ill.zipcode = %s
                   AND ill.source = %s
                   AND NOT (
                       ill.dec_latitude = %s
                   AND ill.dec_longitude = %s
                   );
                """,
                    (
                        io_glob.IO_LAT_LNG_TYPE_ZIPCODE,
                        io_glob.COUNTRY_USA,
                        state.rstrip(),
                        city.rstrip(),
                        zipcode.rstrip(),
                        lat,
                        lng,
                        io_glob.SOURCE_ZCO_ZIP_CODES,
                        datetime.now(),
                        lat,
                        lng,
                        io_glob.SOURCE_ZCO_ZIP_CODES,
                        datetime.now(),
                        io_glob.IO_LAT_LNG_TYPE_ZIPCODE,
                        io_glob.COUNTRY_USA,
                        state.rstrip(),
                        city.rstrip(),
                        zipcode.rstrip(),
                        io_glob.SOURCE_SM_US_CITIES,
                        lat,
                        lng,
                    ),
                )
                count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Execute a query that returns the list of CICTT codes.
# ------------------------------------------------------------------
def _sql_query_cictt_codes(conn_pg: connection) -> list[str]:
    with conn_pg.cursor() as cur:  # type: ignore
        cur.execute(
            """
        SELECT cictt_code
          FROM io_aviation_occurrence_categories;
        """
        )

        data = []

        for row in cur:
            data.append(row[0])

        return sorted(data)


# ------------------------------------------------------------------
# Execute a query that returns the list of US states.
# ------------------------------------------------------------------
def _sql_query_us_states(conn_pg: connection) -> list[str]:
    with conn_pg.cursor() as cur:  # type: ignore
        cur.execute(
            """
        SELECT state
          FROM io_states
         WHERE country = 'USA';
        """
        )

        data = []

        for row in cur:
            data.append(row[0])

        return sorted(data)


# ------------------------------------------------------------------
# Download an MS Access database file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def download_ntsb_msaccess_file(msaccess: str) -> None:
    """Download an MS Access database file.

    Args:
        msaccess (str):
            The MS Access database file without file extension.
    """
    io_glob.logger.debug(io_glob.LOGGER_START)

    filename_zip = msaccess + "." + io_glob.FILE_EXTENSION_ZIP
    url = io_config.settings.download_url_ntsb_prefix + filename_zip

    try:
        file_resp = requests.get(
            url=url,
            allow_redirects=True,
            stream=True,
            timeout=io_config.settings.download_timeout,
        )

        if file_resp.status_code != 200:
            # ERROR.00.906 Unexpected response status code='{status_code}'
            io_utils.terminate_fatal(
                io_glob.ERROR_00_906.replace(
                    "{status_code}", str(file_resp.status_code)
                )
            )

        # INFO.00.013 The connection to the MS Access database file '{msaccess}.zip'
        # on the NTSB download page was successfully established
        io_utils.progress_msg(io_glob.INFO_00_013.replace("{msaccess}", msaccess))

        if not os.path.isdir(io_config.settings.download_work_dir):
            os.makedirs(io_config.settings.download_work_dir)

        filename_zip = os.path.join(
            io_config.settings.download_work_dir.replace("/", os.sep), filename_zip
        )

        no_chunks = 0

        with open(filename_zip, "wb") as file_zip:
            for chunk in file_resp.iter_content(
                chunk_size=io_config.settings.download_chunk_size
            ):
                file_zip.write(chunk)
                no_chunks += 1

        # INFO.00.014 From the file '{msaccess}' {no_chunks} chunks were downloaded
        io_utils.progress_msg(
            io_glob.INFO_00_014.replace("{msaccess}", msaccess).replace(
                "{no_chunks}", str(no_chunks)
            )
        )

        try:
            zipped_files = zipfile.ZipFile(  # pylint: disable=consider-using-with
                filename_zip
            )

            for zipped_file in zipped_files.namelist():
                zipped_files.extract(zipped_file, io_config.settings.download_work_dir)

            zipped_files.close()
        except zipfile.BadZipFile:
            # ERROR.00.907 File'{filename}' is not a zip file
            io_utils.terminate_fatal(
                io_glob.ERROR_00_907.replace("{filename}", filename_zip)
            )

        os.remove(filename_zip)
        # INFO.00.015 The file '{msaccess}.zip'  was successfully unpacked
        io_utils.progress_msg(io_glob.INFO_00_015.replace("{msaccess}", msaccess))

        _check_ddl_changes(msaccess)
    except ConnectionError:
        # ERROR.00.905 Connection problem with url='{url}'
        io_utils.terminate_fatal(io_glob.ERROR_00_905.replace("{url}", url))
    except TimeoutError:
        # ERROR.00.909 Timeout after'{timeout}' seconds with url='{url}
        io_utils.terminate_fatal(
            io_glob.ERROR_00_909.replace(
                "{timeout}", str(io_config.settings.download_timeout)
            ).replace("{url}", url)
        )

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Download a US zip code file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def download_us_cities_file() -> None:
    """Download a US zip code file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    url = io_config.settings.download_url_simplemaps_us_cities

    try:
        file_resp = requests.get(
            url=url,
            allow_redirects=True,
            stream=True,
            timeout=io_config.settings.download_timeout,
        )

        if file_resp.status_code != 200:
            # ERROR.00.906 Unexpected response status code='{status_code}'
            io_utils.terminate_fatal(
                io_glob.ERROR_00_906.replace(
                    "{status_code}", str(file_resp.status_code)
                )
            )

        # INFO.00.030 The connection to the US city file '{filename}'
        # on the simplemaps download page was successfully established
        io_utils.progress_msg(
            io_glob.INFO_00_030.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_cities_zip
            )
        )

        if not os.path.isdir(io_config.settings.download_work_dir):
            os.makedirs(io_config.settings.download_work_dir)

        filename_zip = os.path.join(
            io_config.settings.download_work_dir.replace("/", os.sep),
            io_config.settings.download_file_simplemaps_us_cities_zip,
        )

        no_chunks = 0

        with open(filename_zip, "wb") as file_zip:
            for chunk in file_resp.iter_content(
                chunk_size=io_config.settings.download_chunk_size
            ):
                file_zip.write(chunk)
                no_chunks += 1

        # INFO.00.023 From the file '{filename}' {no_chunks} chunks were downloaded
        io_utils.progress_msg(
            io_glob.INFO_00_023.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_cities_zip
            ).replace("{no_chunks}", str(no_chunks))
        )

        try:
            zipped_files = zipfile.ZipFile(  # pylint: disable=consider-using-with
                filename_zip
            )

            for zipped_file in zipped_files.namelist():
                zipped_files.extract(zipped_file, io_config.settings.download_work_dir)

            zipped_files.close()
        except zipfile.BadZipFile:
            # ERROR.00.907 File '{filename}' is not a zip file
            io_utils.terminate_fatal(
                io_glob.ERROR_00_907.replace("{filename}", filename_zip)
            )

        os.remove(filename_zip)
        # INFO.00.024 The file '{filename}'  was successfully unpacked
        io_utils.progress_msg(
            io_glob.INFO_00_024.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_cities_zip
            )
        )
    except ConnectionError:
        # ERROR.00.905 Connection problem with url='{url}'
        io_utils.terminate_fatal(io_glob.ERROR_00_905.replace("{url}", url))
    except TimeoutError:
        # ERROR.00.909 Timeout after '{timeout}' seconds with url='{url}
        io_utils.terminate_fatal(
            io_glob.ERROR_00_909.replace(
                "{timeout}", str(io_config.settings.download_timeout)
            ).replace("{url}", url)
        )

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Download a US zip code file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def download_us_zips_file() -> None:
    """Download a US zip code file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    url = io_config.settings.download_url_simplemaps_us_zips

    try:
        file_resp = requests.get(
            url=url,
            allow_redirects=True,
            stream=True,
            timeout=io_config.settings.download_timeout,
        )

        if file_resp.status_code != 200:
            # ERROR.00.906 Unexpected response status code='{status_code}'
            io_utils.terminate_fatal(
                io_glob.ERROR_00_906.replace(
                    "{status_code}", str(file_resp.status_code)
                )
            )

        # INFO.00.022 The connection to the US zip code file '{filename}'
        # on the simplemaps download page was successfully established
        io_utils.progress_msg(
            io_glob.INFO_00_022.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_zips_zip
            )
        )

        if not os.path.isdir(io_config.settings.download_work_dir):
            os.makedirs(io_config.settings.download_work_dir)

        filename_zip = os.path.join(
            io_config.settings.download_work_dir.replace("/", os.sep),
            io_config.settings.download_file_simplemaps_us_zips_zip,
        )

        no_chunks = 0

        with open(filename_zip, "wb") as file_zip:
            for chunk in file_resp.iter_content(
                chunk_size=io_config.settings.download_chunk_size
            ):
                file_zip.write(chunk)
                no_chunks += 1

        # INFO.00.023 From the file '{filename}' {no_chunks} chunks were downloaded
        io_utils.progress_msg(
            io_glob.INFO_00_023.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_zips_zip
            ).replace("{no_chunks}", str(no_chunks))
        )

        try:
            zipped_files = zipfile.ZipFile(  # pylint: disable=consider-using-with
                filename_zip
            )

            for zipped_file in zipped_files.namelist():
                zipped_files.extract(zipped_file, io_config.settings.download_work_dir)

            zipped_files.close()
        except zipfile.BadZipFile:
            # ERROR.00.907 File '{filename}' is not a zip file
            io_utils.terminate_fatal(
                io_glob.ERROR_00_907.replace("{filename}", filename_zip)
            )

        os.remove(filename_zip)
        # INFO.00.024 The file '{filename}'  was successfully unpacked
        io_utils.progress_msg(
            io_glob.INFO_00_024.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_zips_zip
            )
        )
    except ConnectionError:
        # ERROR.00.905 Connection problem with url='{url}'
        io_utils.terminate_fatal(io_glob.ERROR_00_905.replace("{url}", url))
    except TimeoutError:
        # ERROR.00.909 Timeout after '{timeout}' seconds with url='{url}
        io_utils.terminate_fatal(
            io_glob.ERROR_00_909.replace(
                "{timeout}", str(io_config.settings.download_timeout)
            ).replace("{url}", url)
        )

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Download the ZIP Code Database file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def download_zip_code_db_file() -> None:
    """Download the ZIP Code Database file."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    url = io_config.settings.download_url_zip_codes_org

    try:
        file_resp = requests.get(
            url=url,
            allow_redirects=True,
            stream=True,
            timeout=io_config.settings.download_timeout,
        )

        if file_resp.status_code != 200:
            # ERROR.00.906 Unexpected response status code='{status_code}'
            io_utils.terminate_fatal(
                io_glob.ERROR_00_906.replace(
                    "{status_code}", str(file_resp.status_code)
                )
            )

        # INFO.00.058 The connection to the Zip Code Database file '{filename}'
        # on the Zip Codes.org download page was successfully established
        io_utils.progress_msg(
            io_glob.INFO_00_058.replace(
                "{filename}", io_config.settings.download_file_zip_codes_org_xlsx
            )
        )

        if not os.path.isdir(io_config.settings.download_work_dir):
            os.makedirs(io_config.settings.download_work_dir)

        filename_xls = os.path.join(
            io_config.settings.download_work_dir.replace("/", os.sep),
            io_config.settings.download_file_zip_codes_org_xlsx,
        )

        no_chunks = 0

        with open(filename_xls, "wb") as file_zip:
            for chunk in file_resp.iter_content(
                chunk_size=io_config.settings.download_chunk_size
            ):
                file_zip.write(chunk)
                no_chunks += 1

        # INFO.00.023 From the file '{filename}' {no_chunks} chunks were downloaded
        io_utils.progress_msg(
            io_glob.INFO_00_023.replace(
                "{filename}", io_config.settings.download_file_zip_codes_org_xlsx
            ).replace("{no_chunks}", str(no_chunks))
        )

    except ConnectionError:
        # ERROR.00.905 Connection problem with url='{url}'
        io_utils.terminate_fatal(io_glob.ERROR_00_905.replace("{url}", url))
    except TimeoutError:
        # ERROR.00.909 Timeout after '{timeout}' seconds with url='{url}
        io_utils.terminate_fatal(
            io_glob.ERROR_00_909.replace(
                "{timeout}", str(io_config.settings.download_timeout)
            ).replace("{url}", url)
        )

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load airports.
# ------------------------------------------------------------------
def load_airport_data() -> None:
    """Load airports."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    _load_airport_data()

    _load_runway_data()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load aviation occurrence categories.
# ------------------------------------------------------------------
def load_aviation_occurrence_categories() -> None:
    """Load aviation occurrence categories."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    _load_aviation_occurrence_categories()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load country and state data.
# ------------------------------------------------------------------
def load_country_state_data() -> None:
    """Load country and state data."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Load the country data.
    # ------------------------------------------------------------------

    # INFO.00.059 Load country data
    io_utils.progress_msg(io_glob.INFO_00_059)
    io_utils.progress_msg("-" * 80)
    _load_country_data(
        conn_pg,
        cur_pg,
    )

    # ------------------------------------------------------------------
    # Load the state data.
    # ------------------------------------------------------------------

    io_utils.progress_msg("-" * 80)

    # INFO.00.060 Load state data
    io_utils.progress_msg(io_glob.INFO_00_060)
    io_utils.progress_msg("-" * 80)
    _load_state_data(
        conn_pg,
        cur_pg,
    )

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    db_utils.upd_io_processed_files(
        io_config.settings.download_file_countries_states_json, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load data from MS Access to the PostgreSQL database.
# ------------------------------------------------------------------
def load_ntsb_msaccess_data(msaccess: str) -> None:
    """Load data from MS Access to the PostgreSQL database.

    Args:
        msaccess (str):
            The MS Access database file without file extension.
    """
    io_glob.logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------

    # pylint: disable=R0801
    filename = os.path.join(
        io_config.settings.download_work_dir.replace("/", os.sep),
        msaccess + "." + io_glob.FILE_EXTENSION_MDB,
    )

    if not os.path.isfile(filename):
        # ERROR.00.912 The MS Access database file '{filename}' is missing
        io_utils.terminate_fatal(io_glob.ERROR_00_912.replace("{filename}", filename))

    conn_ma, cur_ma = db_utils.get_msaccess_cursor(filename)
    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    if msaccess in [io_glob.MSACCESS_AVALL, io_glob.MSACCESS_PRE2008]:
        conn_pg.set_session(autocommit=False)

    if msaccess == io_glob.MSACCESS_PRE2008:
        _delete_ntsb_data(conn_pg, cur_pg)

    is_table_aircraft = False
    is_table_dt_aircraft = False
    is_table_dt_events = False
    is_table_dt_flight_crew = False
    is_table_engines = False
    is_table_events = False
    is_table_events_sequence = False
    is_table_findings = False
    is_table_flight_crew = False
    is_table_flight_time = False
    is_table_injury = False
    is_table_narratives = False
    is_table_ntsb_admin = False
    is_table_occurrences = False
    is_table_seq_of_events = False

    for table_info in cur_ma.tables():
        if table_info.table_name == io_glob.TABLE_NAME_AIRCRAFT:
            is_table_aircraft = True
        elif table_info.table_name == io_glob.TABLE_NAME_DT_AIRCRAFT:
            is_table_dt_aircraft = True
        elif table_info.table_name == io_glob.TABLE_NAME_DT_EVENTS:
            is_table_dt_events = True
        elif table_info.table_name == io_glob.TABLE_NAME_DT_FLIGHT_CREW:
            is_table_dt_flight_crew = True
        elif table_info.table_name == io_glob.TABLE_NAME_ENGINES:
            is_table_engines = True
        elif table_info.table_name == io_glob.TABLE_NAME_EVENTS:
            is_table_events = True
        elif table_info.table_name == io_glob.TABLE_NAME_EVENTS_SEQUENCE:
            is_table_events_sequence = True
        elif table_info.table_name == io_glob.TABLE_NAME_FINDINGS:
            is_table_findings = True
        elif table_info.table_name == io_glob.TABLE_NAME_FLIGHT_CREW:
            is_table_flight_crew = True
        elif table_info.table_name == io_glob.TABLE_NAME_FLIGHT_TIME:
            is_table_flight_time = True
        elif table_info.table_name == io_glob.TABLE_NAME_INJURY:
            is_table_injury = True
        elif table_info.table_name == io_glob.TABLE_NAME_NARRATIVES:
            is_table_narratives = True
        elif table_info.table_name == io_glob.TABLE_NAME_NTSB_ADMIN:
            is_table_ntsb_admin = True
        elif table_info.table_name == io_glob.TABLE_NAME_OCCURRENCES:
            is_table_occurrences = True
        elif table_info.table_name == io_glob.TABLE_NAME_SEQ_OF_EVENTS:
            is_table_seq_of_events = True
        else:
            # INFO.00.021 The following database table is not processed: '{msaccess}'
            io_utils.progress_msg(
                io_glob.INFO_00_021.replace("{msaccess}", table_info.table_name)
            )

    # ------------------------------------------------------------------
    # Load the NTSB data.
    # ------------------------------------------------------------------

    # Level 1 - without FK
    if is_table_events:
        _load_table_events(msaccess, io_glob.TABLE_NAME_EVENTS, cur_ma, conn_pg, cur_pg)

    # Level 2 - FK: ev_id
    if is_table_aircraft:
        _load_table_aircraft(
            msaccess, io_glob.TABLE_NAME_AIRCRAFT, cur_ma, conn_pg, cur_pg
        )
    if is_table_dt_events:
        _load_table_dt_events(
            msaccess, io_glob.TABLE_NAME_DT_EVENTS, cur_ma, conn_pg, cur_pg
        )
    if is_table_ntsb_admin:
        _load_table_ntsb_admin(
            msaccess, io_glob.TABLE_NAME_NTSB_ADMIN, cur_ma, conn_pg, cur_pg
        )

    # Level 3 - FK: ev_id & Aircraft_Key
    if is_table_dt_aircraft:
        _load_table_dt_aircraft(
            msaccess, io_glob.TABLE_NAME_DT_AIRCRAFT, cur_ma, conn_pg, cur_pg
        )
    if is_table_engines:
        _load_table_engines(
            msaccess, io_glob.TABLE_NAME_ENGINES, cur_ma, conn_pg, cur_pg
        )
    if is_table_events_sequence:
        _load_table_events_sequence(
            msaccess, io_glob.TABLE_NAME_EVENTS_SEQUENCE, cur_ma, conn_pg, cur_pg
        )
    if is_table_findings:
        _load_table_findings(
            msaccess, io_glob.TABLE_NAME_FINDINGS, cur_ma, conn_pg, cur_pg
        )
    if is_table_flight_crew:
        _load_table_flight_crew(
            msaccess, io_glob.TABLE_NAME_FLIGHT_CREW, cur_ma, conn_pg, cur_pg
        )
    if is_table_injury:
        _load_table_injury(msaccess, io_glob.TABLE_NAME_INJURY, cur_ma, conn_pg, cur_pg)
    if is_table_narratives:
        _load_table_narratives(
            msaccess, io_glob.TABLE_NAME_NARRATIVES, cur_ma, conn_pg, cur_pg
        )
    if is_table_occurrences:
        _load_table_occurrences(
            msaccess, io_glob.TABLE_NAME_OCCURRENCES, cur_ma, conn_pg, cur_pg
        )

    # Level 4 - FK: ev_id & Aircraft_Key & crew_no
    if is_table_dt_flight_crew:
        _load_table_dt_flight_crew(
            msaccess, io_glob.TABLE_NAME_DT_FLIGHT_CREW, cur_ma, conn_pg, cur_pg
        )
    if is_table_flight_time:
        _load_table_flight_time(
            msaccess, io_glob.TABLE_NAME_FLIGHT_TIME, cur_ma, conn_pg, cur_pg
        )

    # Level 4 - FK: ev_id & Aircraft_Key & Occurrence_No
    if is_table_seq_of_events:
        _load_table_seq_of_events(
            msaccess, io_glob.TABLE_NAME_SEQ_OF_EVENTS, cur_ma, conn_pg, cur_pg
        )

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    conn_pg.set_session(autocommit=True)

    # pylint: disable=R0801
    db_utils.upd_io_processed_files(msaccess, cur_pg)

    cur_pg.close()
    conn_pg.close()
    cur_ma.close()
    conn_ma.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load sequence of events sequence data.
# ------------------------------------------------------------------
def load_sequence_of_events() -> None:
    """Load sequence of events sequence data."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    _load_sequence_of_events()

    # pylint: disable=R0801
    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load simplemaps data.
# ------------------------------------------------------------------
def load_simplemaps_data() -> None:
    """Load simplemaps data."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------
    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Delete existing data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_lat_lng
    WHERE SOURCE IN (%s, %s);
    """,
        (
            io_glob.SOURCE_SM_US_CITIES,
            io_glob.SOURCE_SM_US_ZIP_CODES,
        ),
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()

        # INFO.00.063 Processed data source '{data_source}'
        io_utils.progress_msg(
            io_glob.INFO_00_063.replace("{data_source}", io_glob.SOURCE_SM_US_CITIES)
        )
        io_utils.progress_msg(
            io_glob.INFO_00_063.replace("{data_source}", io_glob.SOURCE_SM_US_ZIP_CODES)
        )
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")
        io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load zip data from a US city file.
    # ------------------------------------------------------------------
    _load_simplemaps_data_zips_from_us_cities(conn_pg, cur_pg)

    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load zip data from a US zip code file.
    # ------------------------------------------------------------------
    _load_simplemaps_data_zips_from_us_zips(conn_pg, cur_pg)

    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load city data from a US city file.
    # ------------------------------------------------------------------
    _load_simplemaps_data_cities_from_us_cities(conn_pg, cur_pg)

    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Delete and insert averaged data.
    # ------------------------------------------------------------------
    _load_table_io_lat_lng_average(conn_pg, cur_pg)

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------
    cur_pg.close()
    conn_pg.close()

    io_glob.logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load ZIP Code Database data.
# ------------------------------------------------------------------
def load_zip_codes_org_data() -> None:
    """Load ZIP Code Database data."""
    io_glob.logger.debug(io_glob.LOGGER_START)

    _load_zip_codes_org_data()

    io_glob.logger.debug(io_glob.LOGGER_END)
