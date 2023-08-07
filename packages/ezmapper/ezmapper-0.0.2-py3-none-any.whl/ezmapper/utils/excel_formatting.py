from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def color_cells(
    excel_file_name: str | Path,
    target_sheet: str,
    reference_sheet: str,
    hex_color: str,
    exclude_column: str,
):
    """
    Colors cells in the target sheet of an Excel file based on the cells in the
    reference sheet.

    This function reads an existing Excel file and applies the specified color
    to the cells in the target sheet if the corresponding cells in the reference
    sheet are not empty. A specific column can be excluded from the coloring
    process.

    Args:
        excel_file_name (str): The name of the Excel file.
        target_sheet (str): The name of the sheet where cells will be colored.
        reference_sheet(str): The name of the sheet used as a reference for
            coloring cells.
        hex_color (str): The hex color code to be applied to the cells in the
            target sheet.
        exclude_column (str): The column letter to be excluded
            from the coloring process.

    Example:
        >>> color_cells("example.xlsx", "Sheet1", "Sheet2", "FFFF00", "A")
    """
    # Load the existing Excel file
    book = load_workbook(excel_file_name)

    # Get the target and reference sheets
    target_ws = book[target_sheet]
    reference_ws = book[reference_sheet]

    # Create a fill object with the provided hex color
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # Iterate through the cells in the reference sheet
    for row in reference_ws.iter_rows(
        min_row=2
    ):  # Start from the second row to exclude the header row
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter != exclude_column and cell.value is not None:
                # If the cell is not empty and not in the excluded column, apply
                # the fill
                target_ws[f"{col_letter}{cell.row}"].fill = fill  # type: ignore

    # Save the changes
    book.save(excel_file_name)


def apply_legacy_formatting(output_file: str | Path, sheet_name: str | None = None):
    """
    Applies legacy formatting to an Excel sheet using openpyxl.

    This function applies specific color, font, and number formatting to an
    Excel sheet based on a pre-defined legacy format.

    Args:
        output_file (str | Path): The output file path for the formatted Excel
        file. sheet_name (str | None): The name of the sheet to be formatted.
    """

    # Load the workbook
    wb = openpyxl.load_workbook(output_file)
    worksheets = _wb_to_worksheets(wb, sheet_name)

    for ws_portfolio in worksheets:
        # Define the fill colors
        dark_green_fill = PatternFill(
            start_color="549E39", end_color="549E39", fill_type="solid"
        )
        light_green_fill = PatternFill(
            start_color="8AB833", end_color="8AB833", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        light_blue_fill = PatternFill(
            start_color="4AB5C4", end_color="4AB5C4", fill_type="solid"
        )

        # Define the fonts
        white_font = Font(color="FFFFFF", size=10)
        black_bold_font = Font(color="000000", bold=True, size=10)

        # Apply header color formatting
        for cell in ws_portfolio[1]:  # type: ignore
            cell.font = white_font
            if 1 <= cell.column <= 5:
                cell.fill = dark_green_fill
            elif cell.column == 6:
                cell.fill = yellow_fill
                cell.font = black_bold_font
            elif 7 <= cell.column <= 16:
                cell.fill = light_green_fill
            elif 17 <= cell.column <= 63:
                cell.fill = light_blue_fill

        # Create a dictionary of header names and their positions (1-based
        # index)
        header_ws_portfolio = {}
        for col_idx, cell in enumerate(ws_portfolio[1], start=1):  # type: ignore
            header_ws_portfolio[cell.value] = col_idx

        # Get the last row in the first column
        last_row = ws_portfolio.max_row
        last_col = ws_portfolio.max_column

        # Apply the General number format to all cells
        for row in ws_portfolio.iter_rows(
            min_row=2, max_col=last_col, max_row=last_row
        ):
            for cell in row:
                cell.number_format = "General"

        # Set the date format
        date_format = "dd/mm/yyyy"
        for column_to_format in [
            "Bond_MaturityDate",
            "Bond_ValuationDate",
            "Bond_CallPutDate",
            "Bond_changedate",
        ]:
            for row_idx in range(2, ws_portfolio.max_row + 1):
                ws_portfolio.cell(
                    row=row_idx, column=header_ws_portfolio[column_to_format]
                ).number_format = date_format

    # Save the changes
    wb.save(output_file)


def get_max_cell_length(column: list[Cell]) -> int:
    """
    Function to get the maximum cell length in a column.

    Args:
        column: List[Cell]: List of cells in a column.

    Returns:
        max_length: int: Maximum cell length.
    """
    max_length = 0
    for cell in column:
        if cell.alignment.wrap_text:
            for line in str(cell.value).split("\n"):
                max_length = max(max_length, len(line))
        else:
            max_length = max(max_length, len(str(cell.value)))
    return max_length


def adjust_column_width(worksheet: Worksheet):
    """
    Function to adjust the width of each column in a worksheet.

    Args:
        worksheet: Worksheet: openpyxl Worksheet object.
    """
    for column in worksheet.columns:
        max_length = get_max_cell_length(column)  # type: ignore
        column_width = max_length + 2
        # check if the cell is not a merged cell before trying to access its
        # column_letter attribute
        if not isinstance(column[0], MergedCell):  # type: ignore
            worksheet.column_dimensions[column[0].column_letter].width = column_width  # type: ignore


def auto_fit_width(excel_file: str | Path, sheet_name: str | None = None):
    """
    Automatically fits the column width of an Excel sheet using openpyxl.

    Args:
        excel_file: (str | Path): The output file path for the formatted Excel
        file. sheet_name (str | None): The name of the sheet to be formatted.
                                  If None, all sheets will be formatted.
    """
    wb = openpyxl.load_workbook(excel_file)
    worksheets = _wb_to_worksheets(wb, sheet_name)

    for worksheet in worksheets:
        adjust_column_width(worksheet)

    wb.save(excel_file)


def auto_wrap_text(excel_file: str | Path, sheet_name: str | None = None):
    """
    Automatically wraps the text in all cells of an Excel sheet using openpyxl.

    This function iterates through all cells in the specified sheet and applies
    the wrap_text alignment property to ensure that the content within each cell
    is wrapped.

    Args:
        excel_file: (str | Path): The input file path for the Excel file.
        sheet_name (str): The name of the sheet to apply the auto-wrap text to.
    """

    # Load the existing Excel file
    book = openpyxl.load_workbook(excel_file)
    worksheets = _wb_to_worksheets(book, sheet_name)

    for ws in worksheets:
        # Apply the wrap_text alignment to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Save the changes
    book.save(excel_file)


def _wb_to_worksheets(wb, sheet_name=None):
    if sheet_name is not None:
        return [wb[sheet_name]]

    return [wb[s] for s in wb.sheetnames]
